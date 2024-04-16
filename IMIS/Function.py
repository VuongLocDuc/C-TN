import os
import sys
import pyodbc as odbc
from pyodbc import Error
import pandas as pd
import numpy as np
from scipy.stats import norm
from functools import reduce
import math
from fpdf import FPDF
import pytz
from datetime import datetime,date
from tkinter import filedialog
import openpyxl
from PyQt5 import QtWidgets,QtCore
from PyQt5.QtCore import QObject,pyqtSignal
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtWidgets import QTableWidget, QHeaderView, QApplication, QTableWidgetItem
from PyQt5.QtGui import QColor
import matplotlib
import matplotlib.pyplot as plt
matplotlib.use('Qt5Agg')
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg
from matplotlib.figure import Figure
from main_login import MainWindow
import json
import base64
import requests
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad
import inventorize3 as inv
# library AI
from math import sqrt
from numpy import array
from numpy import mean
from numpy import std
from pandas import DataFrame
from pandas import concat
from sklearn.metrics import mean_squared_error
from keras.models import Sequential
from keras.layers import LSTM, Dense, Bidirectional
import matplotlib.pyplot as plt
from sklearn.preprocessing import StandardScaler
from scipy.signal import detrend
import pickle

class AppFunction():
    AES_IV = 'bscKHn8REOJ2aikS'
    AES_KEY = 'f51ea9d74f14e0396c7deb66ae1b2c25'
    SCM_USER = 'INF.MAPNET'
    SCM_API = 'http://ftelscmapi.fpt.vn/api/ApiGateway/Post'
    """ docstring for AppFuction"""
    eoq_mh=[]
    eoq_th=[]
    eoq_q=[]
    def __init__(self):
        super(AppFunction,self).__init__()
        # self.arg=arg
    # def combobox_intertwine(self):
    #     if self.ui.loaiKhoCb.currentText() != "Kho tạm":
    #         self.ui.loaiKTCb.setEnabled(False)
    #     else:
    #         self.ui.loaiKTCb.setEnabled(True)
    def Logout(self):
        self.login = MainWindow()  # Pass the data to the second window
        self.hide()
        self.login.show()
    # create database connection
    def create_connection(self):
        DRIVER_NAME='SQL Server'
        SERVER_NAME='DESKTOP-TF6BQMV\SQLEXPRESS01'
        DATABASE_NAME='SCM'
        connection_string= f"""
        DRIVER={{{DRIVER_NAME}}};
        SERVER={SERVER_NAME};
        DATABASE={DATABASE_NAME};
        Trust_Connection=yes;
        """
        try:
            conn=odbc.connect(connection_string)
        except Error as e:
            print(e)
        return conn
    # Bảng kho
    def FilterKho(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        global dfKhoINF
        vung=self.ui.vungCb.currentText()
        cn=self.ui.chiNhanhCb.currentText()
        lk=self.ui.loaiKhoCb.currentText()
        mk=self.ui.maKho.toPlainText()
        tk=self.ui.tenKho.toPlainText()
        c.execute("select MaKho,TenKho,TenCN,TenLoaiKho,TenVung from Vung,Kho,LoaiKho,CN where Vung.MaVung=CN.MaVung and Kho.MaLoaiKho=LoaiKho.MaLoaiKho and Kho.MaCN=CN.MaCN")
        dfKhoINF=c.fetchall()
        data =[]
        for i in dfKhoINF:
            i=tuple(i)
            data.append(i)
        names = [ x[0] for x in c.description]
        dfKhoINF = pd.DataFrame(data, columns=names)
        conditions=[]
        if vung !='Vùng':
            conditions.append(dfKhoINF["TenVung"]==vung)
        if cn !='Chi nhánh':
            conditions.append(dfKhoINF["TenCN"]==cn)
        if lk !='Loại kho':
            conditions.append(dfKhoINF["TenLoaiKho"]==lk)
        # if lkt !='Loại kho tạm':
        #     conditions.append(dfKhoINF["LoaiKT"]==lkt)
        if len(mk)>0: # mk is not null but len =0
            conditions.append(dfKhoINF["MaKho"]==mk)
        if len(tk)>0: # mk is not null but len =0
            conditions.append(dfKhoINF["TenKho"]==tk)
        if len(conditions)>0:
            mask=reduce(lambda x,y: x&y,conditions)
            dfKhoINF=dfKhoINF[mask]
        else:
            dfKhoINF=dfKhoINF
        dfKhoINF=dfKhoINF[['TenVung','TenCN','MaKho','TenKho','TenLoaiKho']]
        if len(dfKhoINF)==0:
            QMessageBox.information(self, "Thông báo", f"Không có kho thỏa mãn!")
        # self.ui.tbKho.clearContents()
        
        self.ui.tbKho.setRowCount(len(dfKhoINF))
        for row_number, (_,row_data) in enumerate(dfKhoINF.iterrows()):
            # self.ui.tbKho.insertRow(row_number)
            for column_number, data in enumerate(row_data):
                self.ui.tbKho.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        self.ui.tbKho.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
        self.ui.tbKho.setColumnWidth(0, 120)
        # Set the width of the second column to 200 pixels
        self.ui.tbKho.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
        self.ui.tbKho.setColumnWidth(1, 170)
        # Set the width of the third column to fill the remaining space
        self.ui.tbKho.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
        self.ui.tbKho.setColumnWidth(2, 140)
        self.ui.tbKho.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
        self.ui.tbKho.setColumnWidth(3, 300)
        self.ui.tbKho.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.Fixed)
        self.ui.tbKho.setColumnWidth(4, 200)
    def addKho(self):
        try:
            conn=AppFunction.create_connection(self)
            c=conn.cursor()
            mk=self.ui.maKho.toPlainText()
            tk=self.ui.tenKho.toPlainText()
            cn=self.ui.chiNhanhCb.currentText()
            lk=self.ui.loaiKhoCb.currentText()
            # # component=["FOX",tentatcn,"INF",lk,]
            # tk="-".join(i for i in component)
            if len(mk)>0 and len(tk)>0 and cn!="Chi nhánh" and lk!="Loại kho": 
                c.execute("select * from Kho where MaKho='"+mk+"'")
                exist=c.fetchone()
                if exist:
                    QMessageBox.information(self, "Thông báo", f"Kho đã tồn tại!")
                else:
                    c.execute("select MaCN from CN where TenCN =N'"+cn+"'")
                    mcn=c.fetchone()[0]
                    c.execute("select MaLoaiKho from LoaiKho where TenLoaiKho=N'"+lk+"'")
                    mlk=c.fetchone()[0]
                    query_string = "insert into Kho (MaKho, TenKho,MaLoaiKho,MaCN) values (?, ?, ?,?)"
                    values = (mk,tk,mlk,mcn)
                    c.execute(query_string, values)
                    conn.commit()
                    QMessageBox.information(self, "Thông báo", f"Thêm kho thành công!")
            else:
                QMessageBox.information(self, "Thông báo", f"Vui lòng nhập đủ thông tin!")
            # clear form input
            mk=self.ui.maKho.setText("")
            tk=self.ui.tenKho.setText("")
            cn=self.ui.chiNhanhCb.setCurrentIndex(0)
            lk=self.ui.loaiKhoCb.setCurrentIndex(0)
            # load new user from db to table view
            AppFunction.displayKho(self)
        except Error as e:
            QMessageBox.information(self, "Thông báo", f"Thêm kho thất bại!")
    def SuaKho(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        mk=self.ui.maKho.toPlainText()
        tk=self.ui.tenKho.toPlainText()
        cn=self.ui.chiNhanhCb.currentText()
        lk=self.ui.loaiKhoCb.currentText()
        c.execute("select MaCN from CN where TenCN =N'"+cn+"'")
        mcn=c.fetchone()[0]
        c.execute("select MaLoaiKho from LoaiKho where TenLoaiKho=N'"+lk+"'")
        mlk=c.fetchone()[0]
        if (lk!="Loại kho") and (len(mk)>0) and (len(tk)>0) and (cn!='Chi nhánh'):
            try:
                c.execute("update Kho set TenKho='{}', MaLoaiKho={}, MaCN='{}' where MaKho='{}'".format(tk, mlk, mcn,mk))
                c.commit()
                QMessageBox.information(self, "Thông báo", f"Sửa thông tin thành công!")
            except:
                QMessageBox.information(self, "Thông báo", f"Sửa thông tin thất bại!")
        else:
            QMessageBox.information(self, "Thông báo", f"Vui lòng nhập đủ thông tin!")
        # c.commit()
        AppFunction.displayKho(self)
        # clear form input
        mk=self.ui.maKho.setText("")
        tk=self.ui.tenKho.setText("")
        cn=self.ui.chiNhanhCb.setCurrentIndex(0)
        lk=self.ui.loaiKhoCb.setCurrentIndex(0)
    def XoaKho(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        row=self.ui.tbKho.currentRow()
        mk=self.ui.tbKho.item(int(row),2).text()
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Question)
        msgBox.setText(f"Bạn có chắc chắn muốn xóa?")
        msgBox.setWindowTitle("Thông báo")
        msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        result = msgBox.exec_()
        if result == QMessageBox.Ok:
            try:
                c.execute("Delete from Kho where MaKho='"+mk+"'")
                c.commit()
                QMessageBox.information(self, "Thông báo", f"Xóa kho thành công!")
            except:
                QMessageBox.information(self, "Thông báo", f"Xóa kho thất bại!")
        AppFunction.displayKho(self)
    def NapKho(self):
        row=self.ui.tbKho.currentRow()
        vung=self.ui.tbKho.item(int(row),0).text()
        cn=self.ui.tbKho.item(int(row),1).text()
        mk=self.ui.tbKho.item(int(row),2).text()
        tk=self.ui.tbKho.item(int(row),3).text()
        lk=self.ui.tbKho.item(int(row),4).text()
        self.ui.vungCb.setCurrentText(vung)
        self.ui.chiNhanhCb.setCurrentText(cn)
        self.ui.maKho.setText(mk)
        self.ui.tenKho.setText(tk)
        if lk=='Kho hàng':
            self.ui.loaiKhoCb.setCurrentIndex(1)
        elif lk=='Kho tạm Triển khai':
            self.ui.loaiKhoCb.setCurrentIndex(2)
        elif lk=='Kho tạm Đối tác':
            self.ui.loaiKhoCb.setCurrentIndex(3)
        elif lk=='Kho tạm DPXLSC':
            self.ui.loaiKhoCb.setCurrentIndex(4)
        elif lk=='Kho trung chuyển':
            self.ui.loaiKhoCb.setCurrentIndex(5)
        elif lk=='Kho CCDC':
            self.ui.loaiKhoCb.setCurrentIndex(6)
        elif lk=='Kho tài sản':
            self.ui.loaiKhoCb.setCurrentIndex(7)
        # if lk!="Kho tạm":
        #     self.ui.loaiKTCb.setCurrentIndex(0)
        # else:
        #     self.ui.loaiKTCb.setCurrentText(lkt)
    def displayKho(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        c.execute("select TenVung,TenCN,MaKho,TenKho,TenLoaiKho from Vung,Kho,LoaiKho,CN where Vung.MaVung=CN.MaVung and Kho.MaLoaiKho=LoaiKho.MaLoaiKho and Kho.MaCN=CN.MaCN")
        dfKhoINF=c.fetchall()
        data =[]
        for i in dfKhoINF:
            i=tuple(i)
            data.append(i)
        names = [ x[0] for x in c.description]
        dfKhoINF = pd.DataFrame(data, columns=names)
        self.ui.tbKho.setRowCount(len(dfKhoINF))
        for row_number, (_,row_data) in enumerate(dfKhoINF.iterrows()):
            # self.ui.tbKho.insertRow(row_number)
            for column_number, data in enumerate(row_data):
                self.ui.tbKho.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        self.ui.tbKho.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
        self.ui.tbKho.setColumnWidth(0, 120)
        # Set the width of the second column to 200 pixels
        self.ui.tbKho.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
        self.ui.tbKho.setColumnWidth(1, 170)
        # Set the width of the third column to fill the remaining space
        self.ui.tbKho.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
        self.ui.tbKho.setColumnWidth(2, 140)
        self.ui.tbKho.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
        self.ui.tbKho.setColumnWidth(3, 300)
        self.ui.tbKho.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.Fixed)
        self.ui.tbKho.setColumnWidth(4, 200)
        # clear form input
        mk=self.ui.maKho.setText("")
        tk=self.ui.tenKho.setText("")
        cn=self.ui.chiNhanhCb.setCurrentIndex(0)
        lk=self.ui.loaiKhoCb.setCurrentIndex(0)
    def BCKho(self):
        global dfKhoINF
        df=dfKhoINF
        file_path = filedialog.asksaveasfilename(defaultextension='.xlsx')
        if file_path:
            # writer = pd.ExcelWriter(file_path,engine="openpyxl")
            df.to_excel(file_path, index=False)
            # writer.save()

    # Kiểm tra tồn kho
    def DisplayB048(self):
        tz_VN = pytz.timezone('Asia/Ho_Chi_Minh')
        datetime_VN = datetime.now(tz_VN)
        now_vn = datetime_VN.strftime("%d/%m/%Y")
        now_vn=pd.to_datetime(now_vn)
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        c.execute("with Kho_CN as (select MaKho,TenKho,TenCN,TenVung from Kho,CN,Vung where Kho.MaCN=CN.MaCN and CN.MaVung=Vung.MaVung) select TenVung,TenCN,TenKho,MaHang,sum(TongTon) as TongTon from B048, Kho_CN where Kho_CN.MaKho=B048.MaKho group by TenVung,TenCN,TenKho,MaHang order by sum(TongTon) desc")
        dfB048=c.fetchall()
        data =[]
        for i in dfB048:
            i=tuple(i)
            data.append(i)
        names = [ x[0] for x in c.description]
        dfB048 = pd.DataFrame(data, columns=names)
        # Lọc dataframe lấy data hôm nay
        # dfB048['ThoiGian']=pd.to_datetime(dfB048['ThoiGian'])
        # dfB048=dfB048[dfB048['ThoiGian']==now_vn]
        c.execute("select MaHang,TenHang from HangHoa")
        dfHH=c.fetchall()
        data1 =[]
        for i in dfHH:
            i=tuple(i)
            data1.append(i)
        names1 = [ x[0] for x in c.description]
        dfHH = pd.DataFrame(data1, columns=names1)
        df=pd.merge(dfB048,dfHH,on="MaHang",how="inner")
        # df.drop(columns="MaHang",inplace=True)
        c.execute("select MaHang,SLYC from A040Detail")
        dfYCGH=c.fetchall()
        data2 =[]
        for i in dfYCGH:
            i=tuple(i)
            data2.append(i)
        names2 = [ x[0] for x in c.description]
        dfYCGH = pd.DataFrame(data2, columns=names2)
        df=df.merge(dfYCGH,on='MaHang',how='left')
        df=df.fillna(0)
        df['SLTonKD']=df['TongTon']-df['SLYC']
        df=df[["TenVung","TenCN","TenKho","TenHang","SLTonKD","TongTon"]]
        if self.ui.tkdRadioBtn.isChecked():
            self.ui.tbCheckTon.setRowCount(len(df))
            for row_number, (_,row_data) in enumerate(df.iterrows()):
                # self.ui.tbKho.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.ui.tbCheckTon.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))   
            self.ui.tbCheckTon.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
            self.ui.tbCheckTon.setColumnWidth(0, 150)
            # Set the width of the second column to 200 pixels
            self.ui.tbCheckTon.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
            self.ui.tbCheckTon.setColumnWidth(1, 170)
            self.ui.tbCheckTon.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
            self.ui.tbCheckTon.setColumnWidth(2, 350)
            self.ui.tbCheckTon.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
            self.ui.tbCheckTon.setColumnWidth(3, 350)
            self.ui.tbCheckTon.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.Fixed)
            self.ui.tbCheckTon.setColumnWidth(4, 210)
            # Set the width of the third column to fill the remaining space
            self.ui.tbCheckTon.horizontalHeader().setSectionResizeMode(5, QtWidgets.QHeaderView.Fixed)
            self.ui.tbCheckTon.setColumnWidth(5, 180)
            delegate = AlignDelegate()
            self.ui.tbCheckTon.setItemDelegateForColumn(4, delegate)
            self.ui.tbCheckTon.setItemDelegateForColumn(5, delegate)
    def DisplayA010(self):
        # A010
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        c.execute("select TenNhomNH,A010Detail.MaHang,TenHang,sum(SLChuaGiao) as SLChuaGiao from A010Detail, NhomNH,HangHoa where NhomNH.MaNhomNH=HangHoa.MaNhomNH and HangHoa.MaHang=A010Detail.MaHang group by TenNhomNH,A010Detail.MaHang,TenHang order by SLChuaGiao desc")
        dfDH=c.fetchall()
        data2 =[]
        for i in dfDH:
            i=tuple(i)
            data2.append(i)
        names2 = [ x[0] for x in c.description]
        dfDH = pd.DataFrame(data2, columns=names2)
        self.ui.dhTb.setRowCount(len(dfDH))
        for row_number, (_,row_data) in enumerate(dfDH.iterrows()):
            for column_number, data in enumerate(row_data):
                self.ui.dhTb.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))   
        self.ui.dhTb.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
        self.ui.dhTb.setColumnWidth(0, 270)
        self.ui.dhTb.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
        self.ui.dhTb.setColumnWidth(1, 200)
        self.ui.dhTb.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
        self.ui.dhTb.setColumnWidth(2, 640)
        self.ui.dhTb.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
        self.ui.dhTb.setColumnWidth(3, 200)
        delegate = AlignDelegate()
        self.ui.dhTb.setItemDelegateForColumn(3, delegate)
    def DisplayA040(self):
        # A040
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        c.execute("with Kho_CN as (select MaKho,TenKho,TenCN,TenVung from Kho,CN,Vung where Kho.MaCN=CN.MaCN and CN.MaVung=Vung.MaVung) select TenVung,TenCN,TenKho,TenHang,sum(SLYC) as SLYC from A040,A040Detail, Kho_CN,HangHoa where Kho_CN.MaKho=A040.MaKhoNhap and HangHoa.MaHang=A040Detail.MaHang group by TenVung,TenCN,TenKho,TenHang order by sum(SLYC) desc")
        dfYCGH=c.fetchall()
        data3 =[]
        for i in dfYCGH:
            i=tuple(i)
            data3.append(i)
        names3 = [ x[0] for x in c.description]
        dfYCGH = pd.DataFrame(data3, columns=names3)
        self.ui.ycghTb.setRowCount(len(dfYCGH))
        for row_number, (_,row_data) in enumerate(dfYCGH.iterrows()):
            for column_number, data in enumerate(row_data):
                self.ui.ycghTb.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))   
        self.ui.ycghTb.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
        self.ui.ycghTb.setColumnWidth(0, 150)
        self.ui.ycghTb.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
        self.ui.ycghTb.setColumnWidth(1, 180)
        self.ui.ycghTb.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
        self.ui.ycghTb.setColumnWidth(2, 400)
        self.ui.ycghTb.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
        self.ui.ycghTb.setColumnWidth(3, 450)
        self.ui.ycghTb.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.Fixed)
        self.ui.ycghTb.setColumnWidth(4, 200)
        delegate = AlignDelegate()
        self.ui.ycghTb.setItemDelegateForColumn(4, delegate)
    def FilterB048(self):
        mh=self.ui.maHangCheckTonFilter.toPlainText()
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        if self.ui.tkdRadioBtn.isChecked():
            c.execute("with Kho_CN as (select MaKho,TenKho,TenCN,TenVung from Kho,CN,Vung where Kho.MaCN=CN.MaCN and CN.MaVung=Vung.MaVung) select TenVung,TenCN,TenKho,MaHang,sum(TongTon) as TongTon from B048, Kho_CN where Kho_CN.MaKho=B048.MaKho group by TenVung,TenCN,TenKho,MaHang order by sum(TongTon) desc")
            dfB048=c.fetchall()
            data =[]
            for i in dfB048:
                i=tuple(i)
                data.append(i)
            names = [ x[0] for x in c.description]
            dfB048 = pd.DataFrame(data, columns=names)
            if len(mh)>0:
                c.execute("select MaHang,TenHang from HangHoa where MaHang='"+mh+"'")
            else:
                c.execute("select MaHang,TenHang from HangHoa")
            dfHH=c.fetchall()
            data1 =[]
            for i in dfHH:
                i=tuple(i)
                data1.append(i)
            names1 = [ x[0] for x in c.description]
            dfHH = pd.DataFrame(data1, columns=names1)
            df=pd.merge(dfB048,dfHH,on="MaHang",how="inner")
            c.execute("select MaHang,SLYC from A040Detail")
            dfYCGH=c.fetchall()
            data2 =[]
            for i in dfYCGH:
                i=tuple(i)
                data2.append(i)
            names2 = [ x[0] for x in c.description]
            dfYCGH = pd.DataFrame(data2, columns=names2)
            df=df.merge(dfYCGH,on='MaHang',how='left')
            df=df.fillna(0)
            df['SLTonKD']=df['TongTon']-df['SLYC']
            df=df[df['SLTonKD']>0]
            df=df[["TenVung","TenCN","TenKho","TenHang","SLTonKD","TongTon"]]
            self.ui.tbCheckTon.setRowCount(len(df))
            for row_number, (_,row_data) in enumerate(df.iterrows()):
                # self.ui.tbKho.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.ui.tbCheckTon.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))   
            self.ui.tbCheckTon.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
            self.ui.tbCheckTon.setColumnWidth(0, 150)
            # Set the width of the second column to 200 pixels
            self.ui.tbCheckTon.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
            self.ui.tbCheckTon.setColumnWidth(1, 350)
            self.ui.tbCheckTon.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
            self.ui.tbCheckTon.setColumnWidth(2, 250)
            self.ui.tbCheckTon.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
            self.ui.tbCheckTon.setColumnWidth(3, 200)
            self.ui.tbCheckTon.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.Fixed)
            self.ui.tbCheckTon.setColumnWidth(4, 200)
            self.ui.tbCheckTon.horizontalHeader().setSectionResizeMode(5, QtWidgets.QHeaderView.Fixed)
            self.ui.tbCheckTon.setColumnWidth(5, 200)
            delegate = AlignDelegate()
            self.ui.tbCheckTon.setItemDelegateForColumn(4, delegate)
            self.ui.tbCheckTon.setItemDelegateForColumn(5, delegate)
            return df
        elif self.ui.dhRadioBtn.isChecked():
            if len(mh)>0:
                c.execute("select TenNhomNH,A010Detail.MaHang,TenHang,sum(SLChuaGiao) as SLChuaGiao from A010Detail, NhomNH,HangHoa where NhomNH.MaNhomNH=HangHoa.MaNhomNH and HangHoa.MaHang=A010Detail.MaHang and A010Detail.MaHang='"+mh+"' group by TenNhomNH,A010Detail.MaHang,TenHang order by SLChuaGiao desc")
            else:
                c.execute("select TenNhomNH,A010Detail.MaHang,TenHang,sum(SLChuaGiao) as SLChuaGiao from A010Detail, NhomNH,HangHoa where NhomNH.MaNhomNH=HangHoa.MaNhomNH and HangHoa.MaHang=A010Detail.MaHang group by TenNhomNH,A010Detail.MaHang,TenHang order by SLChuaGiao desc")
            dfDH=c.fetchall()
            data1 =[]
            for i in dfDH:
                i=tuple(i)
                data1.append(i)
            names1 = [ x[0] for x in c.description]
            dfDH = pd.DataFrame(data1, columns=names1)
            self.ui.dhTb.setRowCount(len(dfDH))
            for row_number, (_,row_data) in enumerate(dfDH.iterrows()):
                for column_number, data in enumerate(row_data):
                    self.ui.dhTb.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))   
            self.ui.dhTb.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
            self.ui.dhTb.setColumnWidth(0, 250)
            self.ui.dhTb.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
            self.ui.dhTb.setColumnWidth(1, 200)
            self.ui.dhTb.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
            self.ui.dhTb.setColumnWidth(2, 400)
            self.ui.dhTb.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
            self.ui.dhTb.setColumnWidth(3, 200)
            delegate = AlignDelegate()
            self.ui.dhTb.setItemDelegateForColumn(3, delegate)
            return dfDH
        elif self.ui.ycghRadioBtn.isChecked():
            if len(mh)>0:
                c.execute("with Kho_CN as (select MaKho,TenKho,TenCN,TenVung from Kho,CN,Vung where Kho.MaCN=CN.MaCN and CN.MaVung=Vung.MaVung) select TenVung,TenCN,TenKho,TenHang,sum(SLYC) as SLYC from A040,A040Detail, Kho_CN,HangHoa where Kho_CN.MaKho=A040.MaKhoNhap and HangHoa.MaHang=A040Detail.MaHang and A040Detail.MaHang='"+mh+"' group by TenVung,TenCN,TenKho,TenHang order by sum(SLYC) desc")
            else:
                c.execute("with Kho_CN as (select MaKho,TenKho,TenCN,TenVung from Kho,CN,Vung where Kho.MaCN=CN.MaCN and CN.MaVung=Vung.MaVung) select TenVung,TenCN,TenKho,TenHang,sum(SLYC) as SLYC from A040,A040Detail, Kho_CN,HangHoa where Kho_CN.MaKho=A040.MaKhoNhap and HangHoa.MaHang=A040Detail.MaHang group by TenVung,TenCN,TenKho,TenHang order by sum(SLYC) desc")
            dfYCGH=c.fetchall()
            data2 =[]
            for i in dfYCGH:
                i=tuple(i)
                data2.append(i)
            names2 = [ x[0] for x in c.description]
            dfYCGH = pd.DataFrame(data2, columns=names2)
            self.ui.ycghTb.setRowCount(len(dfYCGH))
            for row_number, (_,row_data) in enumerate(dfYCGH.iterrows()):
                for column_number, data in enumerate(row_data):
                    self.ui.ycghTb.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))   
            self.ui.ycghTb.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
            self.ui.ycghTb.setColumnWidth(0, 200)
            self.ui.ycghTb.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
            self.ui.ycghTb.setColumnWidth(1, 200)
            self.ui.ycghTb.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
            self.ui.ycghTb.setColumnWidth(2, 400)
            self.ui.ycghTb.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
            self.ui.ycghTb.setColumnWidth(3, 300)
            self.ui.ycghTb.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.Fixed)
            self.ui.ycghTb.setColumnWidth(4, 200)
            delegate = AlignDelegate()
            self.ui.ycghTb.setItemDelegateForColumn(4, delegate)
            return dfYCGH
    # def XuatExcelB048(self):
    #     df=AppFunction.FilterB048(self)
    #     file_path = filedialog.asksaveasfilename(defaultextension='.xlsx',initialfile="B048")
    #     if file_path:
    #         writer = pd.ExcelWriter(file_path,engine="openpyxl")
    #         df.to_excel(writer, index=False)
    def XuatReportB048(self):
        df=AppFunction.FilterB048(self)
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        th=df['TenHang'][0]
        c.execute("select MaHang from HangHoa where TenHang like N'"+th+"'")
        kq=c.fetchone()
        mh=kq[0]
        data=df.values.tolist()
        if self.ui.tkdRadioBtn.isChecked():
        # df.rename(columns={'TenVung':'Vùng','TenCN':'Chi nhánh','TenKho':'Kho','TenHang':'Tên hàng','SLTonKD':'SL tồn khả dụng','TongTon':'Tổng tồn'},inplace=True)
            col=['Vùng','Chi nhánh','Kho','Tên hàng','SL tồn khả dụng','Tổng tồn']
        # data=df.values.tolist() 
            pdf = PDFB048()
            pdf.add_page()
            pdf.alias_nb_pages()
            pdf.set_auto_page_break(False)
            pdf.set_widths([20, 25, 60,40,22.5,22.5])
            pdf.set_font('DejaVu',size=10)
            pdf.set_x(10)
            pdf.multi_cell(0, 5,'Công ty Cổ phần Viễn thông FPT \nTrung tâm Phát triển và Quản lý hạ tầng MB \nĐịa chỉ: 48 Vạn Bảo, Ngọc Khánh, Ba Đình', border="B", align='L')
            pdf.ln(5) 
            pdf.set_font('DejaVu', 'B', 15)
            pdf.multi_cell(0, 7, 'BÁO CÁO TỒN KHO THEO MÃ HÀNG: '+mh, border=0, align='C')
            pdf.ln(3)
        elif self.ui.dhRadioBtn.isChecked():
            col=['Tên nhóm NH','Mã hàng','Tên hàng','SL chưa giao']
            pdf = PDFA010()
            pdf.add_page()
            pdf.alias_nb_pages()
            pdf.set_auto_page_break(False)
            pdf.set_widths([30, 40, 70,50]) 
            pdf.set_font('DejaVu',size=10)
            pdf.set_x(10)
            pdf.multi_cell(0, 5,'Công ty Cổ phần Viễn thông FPT \nTrung tâm Phát triển và Quản lý hạ tầng MB \nĐịa chỉ: 48 Vạn Bảo, Ngọc Khánh, Ba Đình', border="B", align='L')
            pdf.ln(5) 
            pdf.set_font('DejaVu', 'B', 15)
            pdf.multi_cell(0, 7, 'BÁO CÁO TỒN ĐƠN HÀNG THEO MÃ HÀNG: '+mh, border=0, align='C')
            pdf.ln(3)
        elif self.ui.ycghRadioBtn.isChecked():
            col=['Vùng','Chi nhánh','Kho nhập','Tên hàng','SL chờ duyệt']
            pdf = PDFA040()
            pdf.add_page()
            pdf.alias_nb_pages()
            pdf.set_auto_page_break(False)
            pdf.set_widths([20, 25, 60,60,25])
            pdf.set_font('DejaVu',size=10)
            pdf.set_x(10)
            pdf.multi_cell(0, 5,'Công ty Cổ phần Viễn thông FPT \nTrung tâm Phát triển và Quản lý hạ tầng MB \nĐịa chỉ: 48 Vạn Bảo, Ngọc Khánh, Ba Đình', border="B", align='L')
            pdf.ln(5) 
            pdf.set_font('DejaVu', 'B', 15)
            pdf.multi_cell(0, 7, 'BÁO CÁO TỒN YCGH THEO MÃ HÀNG: '+mh, border=0, align='C')
            pdf.ln(3)
        pdf.set_font('DejaVu', '', 10)
        tz_VN = pytz.timezone('Asia/Ho_Chi_Minh')
        datetime_VN = datetime.now(tz_VN)
        now_vn = datetime_VN.strftime("%d/%m/%Y")
        pdf.multi_cell(0, 7, f'Ngày: {now_vn}', border=0, align='R')
        pdf.ln(5)
        pdf.tieude(col) 
        for i in range(len(data)):
            if pdf.get_y() + 30 > pdf.h:
                pdf.add_page()
                pdf.tieude(col)
            pdf.row([data[i]])
        if self.ui.tkdRadioBtn.isChecked():
            sum1=np.round(df["SLTonKD"].sum(),0)
            sum2=np.round(df["TongTon"].sum(),0)
            tong=["Tổng cộng",sum1,sum2]
            pdf.sum(tong)
        elif self.ui.dhRadioBtn.isChecked():
            sum1=np.round(df["SLChuaGiao"].sum(),0)
            tong=["Tổng cộng",sum1]
            pdf.sum(tong)
        elif self.ui.ycghRadioBtn.isChecked():
            sum1=np.round(df["SLYC"].sum(),0)
            tong=["Tổng cộng",sum1]
            pdf.sum(tong)
        # pdf.row(data)
        filename = filedialog.asksaveasfilename(defaultextension='.pdf')
        if filename:
            pdf.output(filename, 'F')
    def DeleteData(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        tz_VN = pytz.timezone('Asia/Ho_Chi_Minh')
        datetime_VN = datetime.now(tz_VN)
        now_vn = datetime_VN.strftime("%d%m%y")
        namtc = int(datetime_VN.strftime("%Y"))-1
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Question)
        msgBox.setText(f"Bạn có chắc chắn muốn xóa dữ liệu năm tài chính {namtc}")
        msgBox.setWindowTitle("Thông báo")
        msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        result = msgBox.exec_()
        if result == QMessageBox.Ok:
            try:
                c.execute("delete from A010Detail where MaDH in (select MaDH from A010 where NamTC ='"+namtc+"')")
                c.execute("delete from A010 where NamTC =?",(namtc,))
                c.execute("delete from A040Detail where MaYCGH in (select MaYCGH from A040 where NamTC ='"+namtc+"')")
                c.execute("delete from A040 where NamTC =?",(namtc,))
                c.commit()
                QMessageBox.information(self, "Thông báo", f"Xóa dữ liệu thành công!")
            except:
                QMessageBox.information(self, "Thông báo", f"Xảy ra lỗi trong quá trình xóa dữ liệu!")
    # Bảng tài khoản
    def displayTK(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        c.execute("select * from TK")
        dfTK=c.fetchall()
        data =[]
        for i in dfTK:
            i=tuple(i)
            data.append(i)
        names = [ x[0] for x in c.description]
        dfTK = pd.DataFrame(data, columns=names)
        dfTK['LoaiTK']=np.where(dfTK['LoaiTK']==1,'Admin','Nhân viên')
        self.ui.tbTaiKhoan.setRowCount(len(dfTK))
        for row_number, (_,row_data) in enumerate(dfTK.iterrows()):
            for column_number, data in enumerate(row_data):
                self.ui.tbTaiKhoan.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        self.ui.tbTaiKhoan.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
        self.ui.tbTaiKhoan.setColumnWidth(0, 180)
        # Set the width of the second column to 200 pixels
        self.ui.tbTaiKhoan.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
        self.ui.tbTaiKhoan.setColumnWidth(1, 180)
        # Set the width of the third column to fill the remaining space
        self.ui.tbTaiKhoan.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
        self.ui.tbTaiKhoan.setColumnWidth(2, 240)
        delegate = AlignRDelegate()
        for i in range(len(dfTK.columns)):
            self.ui.tbTaiKhoan.setItemDelegateForColumn(i, delegate)
        # clear form input
        self.ui.tenTK.setText("")
        self.ui.matKhau.setText("")
        self.ui.loaiTK.setCurrentIndex(0)
    def FilterTK(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        global dfTK
        ltk=self.ui.loaiTK.currentText()
        ttk=self.ui.tenTK.toPlainText()
        c.execute("Select * from TK")
        dfTK=c.fetchall()
        data =[]
        for i in dfTK:
            i=tuple(i)
            data.append(i)
        names = [ x[0] for x in c.description]
        dfTK = pd.DataFrame(data, columns=names)
        if len(ttk)>0 :
            dfTK=dfTK[dfTK["TenTK"]==str(ttk).strip()]
        if ltk !='Loại tài khoản':
            if str(ltk)=="Admin":
                dfTK=dfTK[dfTK["LoaiTK"]==1]
            elif str(ltk)=="Nhân viên":
                dfTK=dfTK[dfTK["LoaiTK"]==0]
        dfTK=dfTK[['TenTK','MK','LoaiTK']]
        dfTK['LoaiTK']=np.where(dfTK['LoaiTK']==1,'Admin','Nhân viên')
        if len(dfTK)==0:
            QMessageBox.information(self, "Thông báo", f"Không có tài khoản thỏa mãn!")
        self.ui.tbTaiKhoan.setRowCount(len(dfTK))
        for row_number, (_,row_data) in enumerate(dfTK.iterrows()):
            for column_number, data in enumerate(row_data):
                self.ui.tbTaiKhoan.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        self.ui.tbTaiKhoan.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
        self.ui.tbTaiKhoan.setColumnWidth(0, 180)
        # Set the width of the second column to 200 pixels
        self.ui.tbTaiKhoan.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
        self.ui.tbTaiKhoan.setColumnWidth(1, 180)
        # Set the width of the third column to fill the remaining space
        self.ui.tbTaiKhoan.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
        self.ui.tbTaiKhoan.setColumnWidth(2, 240)
        delegate = AlignRDelegate()
        for i in range(len(dfTK.columns)):
            self.ui.tbTaiKhoan.setItemDelegateForColumn(i, delegate)
    def ThemTK(self):
        try:
            conn=AppFunction.create_connection(self)
            c=conn.cursor()
            ttk=self.ui.tenTK.toPlainText()
            mk=self.ui.matKhau.toPlainText()
            ltk=self.ui.loaiTK.currentText()
            if len(ttk)>0 and len(mk)>0 and ltk!="Loại tài khoản": 
                c.execute("select * from TK where TenTK='"+ttk+"'")
                exist=c.fetchone()
                if exist:
                    QMessageBox.information(self, "Thông báo", f"Tài khoản đã tồn tại!")
                else:
                    c.execute("select * from NhanVien where MaNV='"+ttk+"'")
                    exist1=c.fetchone()
                    if exist1:
                        if ltk=="Admin":
                            ltk=1
                        else:
                            ltk=0
                        query_string = "insert into TK (TenTK, MK,LoaiTK) values (?, ?,?)"
                        values = (ttk,mk,ltk)
                        c.execute(query_string, values)
                        conn.commit()
                        QMessageBox.information(self, "Thông báo", f"Thêm tài khoản thành công!")
                    else:
                        QMessageBox.information(self, "Thông báo", f"Không tồn tại mã nhân viên!")
            else:
                QMessageBox.information(self, "Thông báo", f"Vui lòng nhập đủ thông tin!")
            # clear form input
            self.ui.tenTK.setText("")
            self.ui.matKhau.setText("")
            self.ui.loaiTK.setCurrentIndex(0)
            # load new user from db to table view
            AppFunction.displayTK(self)
        except Error as e:
            QMessageBox.information(self, "Thông báo", f"Thêm tài khoản thất bại!")
    def SuaTK(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        ttk=self.ui.tenTK.toPlainText()
        mk=self.ui.matKhau.toPlainText()
        ltk=self.ui.loaiTK.currentText()
        if ltk=="Admin":
            ltk=1
        elif ltk=="Nhân viên":
            ltk=0
        else:
            QMessageBox.information(self, "Thông báo", f"Loại tài khoản không tồn tại!")
        try:
            c.execute("update TK set MK='{}', LoaiTK={} where TenTK='{}'".format(mk, ltk, ttk))
            c.commit()
            QMessageBox.information(self, "Thông báo", f"Sửa thông tin thành công!")
        except:
            QMessageBox.information(self, "Thông báo", f"Sửa thông tin thất bại!")
        AppFunction.displayTK(self)
        # clear form input
        self.ui.tenTK.setText("")
        self.ui.matKhau.setText("")
        self.ui.loaiTK.setCurrentIndex(0)
    def XoaTK(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        row=self.ui.tbTaiKhoan.currentRow()
        ttk=self.ui.tbTaiKhoan.item(int(row),0).text()
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Question)
        msgBox.setText(f"Bạn có chắc chắn muốn xóa?")
        msgBox.setWindowTitle("Thông báo")
        msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        result = msgBox.exec_()
        if result == QMessageBox.Ok:
            try:
                c.execute("Delete from TK where TenTK='"+ttk+"'")
                c.commit()
                QMessageBox.information(self, "Thông báo", f"Xóa tài khoản thành công!")
            except:
                QMessageBox.information(self, "Thông báo", f"Xóa tài khoản thất bại!")
        AppFunction.displayTK(self)
    def NapTK(self):
        row=self.ui.tbTaiKhoan.currentRow()
        ttk=self.ui.tbTaiKhoan.item(int(row),0).text()
        mk=self.ui.tbTaiKhoan.item(int(row),1).text()
        ltk=self.ui.tbTaiKhoan.item(int(row),2).text()
        self.ui.tenTK.setText(ttk)
        self.ui.matKhau.setText(mk)
        if ltk=='Admin':
            self.ui.loaiTK.setCurrentIndex(1)
        elif ltk=='Nhân viên':
            self.ui.loaiTK.setCurrentIndex(2)
    # Bảng nhân viên
    def displayNV(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        c.execute("select MaNV,TenNV,email,DiaChi,SDT,NgayVaoLam,TenBPTat from NhanVien,BoPhan where NhanVien.MaBP=BoPhan.MaBP")
        dfNV=c.fetchall()
        data =[]
        for i in dfNV:
            i=tuple(i)
            data.append(i)
        names = [ x[0] for x in c.description]
        dfNV = pd.DataFrame(data, columns=names)
        self.ui.tbNhanVien.setRowCount(len(dfNV))
        for row_number, (_,row_data) in enumerate(dfNV.iterrows()):
            for column_number, data in enumerate(row_data):
                self.ui.tbNhanVien.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        self.ui.tbNhanVien.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
        self.ui.tbNhanVien.setColumnWidth(0, 100)
        # Set the width of the second column to 200 pixels
        self.ui.tbNhanVien.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
        self.ui.tbNhanVien.setColumnWidth(1, 300)
        # Set the width of the third column to fill the remaining space
        self.ui.tbNhanVien.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
        self.ui.tbNhanVien.setColumnWidth(2, 200)
        self.ui.tbNhanVien.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
        self.ui.tbNhanVien.setColumnWidth(3, 300)
        self.ui.tbNhanVien.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.Fixed)
        self.ui.tbNhanVien.setColumnWidth(4, 100)
        self.ui.tbNhanVien.horizontalHeader().setSectionResizeMode(5, QtWidgets.QHeaderView.Fixed)
        self.ui.tbNhanVien.setColumnWidth(5, 150)
        self.ui.tbNhanVien.horizontalHeader().setSectionResizeMode(6, QtWidgets.QHeaderView.Fixed)
        self.ui.tbNhanVien.setColumnWidth(6, 100)
        # self.ui.tbNhanVien.horizontalHeader().setSectionResizeMode(7, QtWidgets.QHeaderView.Stretch)
    def FilterNV(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        mnv=self.ui.maNhanVien.toPlainText()
        tnv=self.ui.tenNhanVien.toPlainText()
        sdt=self.ui.sdtNhanVien.toPlainText()
        email=self.ui.emailNhanVien.toPlainText()
        dc=self.ui.diaChiNhanVien.toPlainText()
        date_temp=self.ui.dateVaoLam.date()
        date=date_temp.toPyDate()
        bp=self.ui.boPhanCb.currentText()
        c.execute("select MaNV,TenNV,email,DiaChi,SDT,NgayVaoLam,TenBPTat from NhanVien,BoPhan where NhanVien.MaBP=BoPhan.MaBP")
        dfNV=c.fetchall()
        data =[]
        for i in dfNV:
            i=tuple(i)
            data.append(i)
        names = [ x[0] for x in c.description]
        dfNV = pd.DataFrame(data, columns=names)
        conditions=[]
        if len(mnv)>0:
            conditions.append(dfNV["MaNV"]==mnv)
        if len(tnv):
            conditions.append(dfNV["TenNV"]==tnv)
        if len(email)>0:
            conditions.append(dfNV["email"]==email)
        if len(dc)>0:
            conditions.append(dfNV["DiaChi"]==dc)
        if len(sdt)>0: # mk is not null but len =0
            conditions.append(dfNV["SDT"]==sdt)
        if date!="2000-01-01":
            conditions.append(dfNV["NgayVaoLam"]==date)
        if bp!="Bộ Phận":
            conditions.append(dfNV["TenBPTat"]==bp)
        mask=reduce(lambda x,y: x&y,conditions)
        dfNV=dfNV[mask]
        self.ui.tbNhanVien.setRowCount(len(dfNV))
        for row_number, (_,row_data) in enumerate(dfNV.iterrows()):
            for column_number, data in enumerate(row_data):
                self.ui.tbNhanVien.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
    def ThemNV(self):
        try:
            conn=AppFunction.create_connection(self)
            c=conn.cursor()
            mnv=self.ui.maNhanVien.toPlainText()
            tnv=self.ui.tenNhanVien.toPlainText()
            sdt=self.ui.sdtNhanVien.toPlainText()
            email=self.ui.emailNhanVien.toPlainText()
            dc=self.ui.diaChiNhanVien.toPlainText()
            date_temp=self.ui.dateVaoLam.date()
            date=date_temp.toPyDate()
            bp=self.ui.boPhanCb.currentText()
            if len(mnv)>0 and len(tnv)>0 and len(sdt)>0 and len(email)>0 and len(dc)>0 and date!="2000-01-01" and bp!="Bộ phận": 
                c.execute("select * from NhanVien where MaNV='"+mnv+"'")
                exist=c.fetchone()
                if exist:
                    msgBox = QMessageBox()
                    msgBox.setIcon(QMessageBox.Information)
                    msgBox.setText("Nhân viên đã tồn tồn tại!")
                    msgBox.setWindowTitle("Thông báo")
                    msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
                    msgBox.exec_()
                else:
                    c.execute("select MaBP from BoPhan where TenBPTat ='"+bp+"'")
                    mbp=c.fetchone()[0]
                    query_string = "insert into NhanVien (MaNV,TenNV,email,DiaChi,SDT,NgayVaoLam,MaBP) values (?, ?,?,?,?,?,?)"
                    values = (mnv,tnv,email,dc,sdt,date,mbp)
                    c.execute(query_string, values)
                    conn.commit()
                    msgBox = QMessageBox()
                    msgBox.setIcon(QMessageBox.Information)
                    msgBox.setText("Thêm nhân viên thành công!")
                    msgBox.setWindowTitle("Thông báo")
                    msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
                    msgBox.exec_()
            else:
                msgBox = QMessageBox()
                msgBox.setIcon(QMessageBox.Information)
                msgBox.setText("Vui lòng nhập đủ thông tin!")
                msgBox.setWindowTitle("Thông báo")
                msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
                msgBox.exec_()
            # clear form input
            self.ui.maNhanVien.setText("")
            self.ui.tenNhanVien.setText("")
            self.ui.emailNhanVien.setText("")
            self.ui.diaChiNhanVien.setText("")
            self.ui.sdtNhanVien.setText("")
            self.ui.dateVaoLam.setDate(self.QDate.currentDate())
            self.ui.boPhanCb.setCurrentIndex(0)
            # load new user from db to table view
            AppFunction.displayNV(self)
        except Error as e:
            msgBox = QMessageBox()
            msgBox.setIcon(QMessageBox.Information)
            msgBox.setText("Thêm nhân viên thất bại!")
            msgBox.setWindowTitle("Thông báo")
            msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            msgBox.exec_()
    def SuaNV(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        mnv=self.ui.maNhanVien.toPlainText()
        tnv=self.ui.tenNhanVien.toPlainText()
        sdt=self.ui.sdtNhanVien.toPlainText()
        email=self.ui.emailNhanVien.toPlainText()
        dc=self.ui.diaChiNhanVien.toPlainText()
        date_temp=self.ui.dateVaoLam.date()
        date=date_temp.toPyDate()
        bp=self.ui.boPhanCb.currentText()
        if len(mnv)>0 and len(tnv)>0 and len(sdt)>0 and len(email)>0 and len(dc)>0 and date!="2000-01-01" and bp!="Bộ phận": 
            c.execute("select MaBP from BoPhan where TenBPTat ='"+bp+"'")
            mbp=c.fetchone()[0]
            # if bp=="Bộ phận":
            #     msgBox = QMessageBox()
            #     msgBox.setIcon(QMessageBox.Information)
            #     msgBox.setText("Bộ phận không tồn tại!")
            #     msgBox.setWindowTitle("Thông báo")
            #     msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            #     msgBox.exec_()
            c.execute("update NhanVien set TenNV='{}', email={},DiaChi='{}',SDT={},NgayVaoLam={},MaBP={} where MaNV={}".format(tnv, email, dc,sdt,date,mbp))
            c.commit()
            AppFunction.displayNV(self)
        else:
            msgBox = QMessageBox()
            msgBox.setIcon(QMessageBox.Information)
            msgBox.setText("Vui lòng nhập đủ thông tin!")
            msgBox.setWindowTitle("Thông báo")
            msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            msgBox.exec_()
    def XoaNV(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        row=self.ui.tbNhanVien.currentRow()
        mnv=self.ui.tbNhanVien.item(int(row),0).text()
        c.execute("Delete from NhanVien where MaNV='"+mnv+"'")
        c.commit()
        AppFunction.displayNV(self)
    def NapNV(self):
        row=self.ui.tbNhanVien.currentRow()
        mnv=self.ui.tbNhanVien.item(int(row),0).text()
        tnv=self.ui.tbNhanVien.item(int(row),1).text()
        email=self.ui.tbNhanVien.item(int(row),2).text()
        dc=self.ui.tbNhanVien.item(int(row),3).text()
        sdt=self.ui.tbNhanVien.item(int(row),4).text()
        nvl=self.ui.tbNhanVien.item(int(row),5).text()
        bp=self.ui.tbNhanVien.item(int(row),6).text()
        self.ui.maNhanVien.setText(mnv)
        self.ui.tenNhanVien.setText(tnv)
        self.ui.sdtNhanVien.setText(sdt)
        self.ui.emailNhanVien.setText(email)
        self.ui.diaChiNhanVien.setText(dc)
        self.ui.dateVaoLam.setDate(nvl)
        self.ui.boPhanCb.setCurrentText(bp)
    def BCHH(self):
        global dfHH
        df=dfHH
        file_path = filedialog.asksaveasfilename(defaultextension='.xlsx')
        if file_path:
            # writer = pd.ExcelWriter(file_path,engine="openpyxl")
            df.to_excel(file_path, index=False)
    # Bảng hàng hóa
    def displayHH(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        c.execute("select MaHang,TenHang,DVT,DonGia,TenNhomNH,LT from HangHoa,NhomNH where HangHoa.MaNhomNH=NhomNH.MaNhomNH")
        dfHH=c.fetchall()
        data =[]
        for i in dfHH:
            i=tuple(i)
            data.append(i)
        names = [ x[0] for x in c.description]
        dfHH = pd.DataFrame(data, columns=names)
        self.ui.tbHangHoa.setRowCount(len(dfHH))
        for row_number, (_,row_data) in enumerate(dfHH.iterrows()):
            for column_number, data in enumerate(row_data):
                self.ui.tbHangHoa.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        self.ui.tbHangHoa.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
        self.ui.tbHangHoa.setColumnWidth(0, 150)
        # Set the width of the third column to fill the remaining space
        self.ui.tbHangHoa.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
        self.ui.tbHangHoa.setColumnWidth(1, 300)
        self.ui.tbHangHoa.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
        self.ui.tbHangHoa.setColumnWidth(2, 100)
        self.ui.tbHangHoa.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
        self.ui.tbHangHoa.setColumnWidth(3, 150)
        self.ui.tbHangHoa.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.Fixed)
        self.ui.tbHangHoa.setColumnWidth(4, 270)
        self.ui.tbHangHoa.horizontalHeader().setSectionResizeMode(5, QtWidgets.QHeaderView.Fixed)
        self.ui.tbHangHoa.setColumnWidth(5, 80)
        # self.ui.tbHangHoa.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.Stretch)
        delegate = AlignDelegate()
        self.ui.tbHangHoa.setItemDelegateForColumn(4, delegate)
        self.ui.tbHangHoa.setItemDelegateForColumn(5, delegate)
        # clear form input
        self.ui.maHang.setText("")
        # self.ui.maHangTD.setText("")
        self.ui.tenHang.setText("")
        self.ui.dvt.setText("")
        self.ui.donGia.setText("")
        self.ui.nhomNH.setCurrentIndex(0)
        self.ui.leadTime.setText("")
    def FilterHH(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        global dfHH
        mh=self.ui.maHang.toPlainText()
        # mhtd=self.ui.maHangTD.toPlainText()
        th=self.ui.tenHang.toPlainText()
        dvt=self.ui.dvt.toPlainText()
        dg=self.ui.donGia.toPlainText()
        nnh=self.ui.nhomNH.currentText()
        lt=self.ui.leadTime.toPlainText()
        c.execute("select MaHang,TenHang,DVT,DonGia,TenNhomNH,LT from HangHoa,NhomNH where HangHoa.MaNhomNH=NhomNH.MaNhomNH")
        dfHH=c.fetchall()
        data =[]
        for i in dfHH:
            i=tuple(i)
            data.append(i)
        names = [ x[0] for x in c.description]
        dfHH = pd.DataFrame(data, columns=names)
        conditions=[]
        if len(mh)>0:
            conditions.append(dfHH["MaHang"]==mh)
        # if len(mhtd):
        #     conditions.append(dfHH["MaHangTD"]==mhtd)
        if len(th)>0:
            conditions.append(dfHH["TenHang"]==th)
        if len(dvt)>0:
            conditions.append(dfHH["DVT"]==dvt)
        if len(dg)>0: # mk is not null but len =0
            conditions.append(dfHH["DonGia"]==dg)
        if nnh!="Nhóm ngành hàng":
            conditions.append(dfHH["TenNhomNH"]==nnh)
        if len(lt):
            conditions.append(dfHH["LT"]==lt)
        if len(conditions)>0:
            mask=reduce(lambda x,y: x&y,conditions)
            dfHH=dfHH[mask]
        else:
            dfHH=dfHH
        dfHH=dfHH[['MaHang','TenHang','DVT','DonGia','TenNhomNH','LT']]
        if len(dfHH)==0:
            QMessageBox.information(self, "Thông báo", f"Không có hàng hóa thỏa mãn!")
        
        self.ui.tbHangHoa.setRowCount(len(dfHH))
        for row_number, (_,row_data) in enumerate(dfHH.iterrows()):
            for column_number, data in enumerate(row_data):
                self.ui.tbHangHoa.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        self.ui.tbHangHoa.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
        self.ui.tbHangHoa.setColumnWidth(0, 150)
        # Set the width of the third column to fill the remaining space
        self.ui.tbHangHoa.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
        self.ui.tbHangHoa.setColumnWidth(1, 300)
        self.ui.tbHangHoa.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
        self.ui.tbHangHoa.setColumnWidth(2, 100)
        self.ui.tbHangHoa.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
        self.ui.tbHangHoa.setColumnWidth(3, 150)
        self.ui.tbHangHoa.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.Fixed)
        self.ui.tbHangHoa.setColumnWidth(4, 270)
        self.ui.tbHangHoa.horizontalHeader().setSectionResizeMode(5, QtWidgets.QHeaderView.Fixed)
        self.ui.tbHangHoa.setColumnWidth(5, 80)
        # self.ui.tbHangHoa.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.Stretch)
        delegate = AlignDelegate()
        self.ui.tbHangHoa.setItemDelegateForColumn(4, delegate)
        self.ui.tbHangHoa.setItemDelegateForColumn(5, delegate)
    def ThemHH(self):
        try:
            conn=AppFunction.create_connection(self)
            c=conn.cursor()
            mh=self.ui.maHang.toPlainText()
            # mhtd=self.ui.maHangTD.toPlainText()
            th=self.ui.tenHang.toPlainText()
            dvt=self.ui.dvt.toPlainText()
            dg=self.ui.donGia.toPlainText()
            nnh=self.ui.nhomNH.currentText()
            lt=self.ui.leadTime.toPlainText()
            if len(mh)>0 and len(th)>0 and len(dvt)>0 and len(dg)>0 and nnh!="Nhóm ngành hàng" and len(lt)>0: 
                c.execute("select * from HangHoa where MaHang='"+mh+"'")
                exist=c.fetchone()
                if exist:
                    QMessageBox.information(self, "Thông báo", f"Hàng hóa đã tồn tại!")
                else:
                    c.execute("select MaNhomNH from NhomNH where TenNhomNH like N'"+nnh+"'")
                    mnnh=c.fetchone()[0]
                    # c.execute("select MaChungLoaiHH from ChungLoaiHH where TenChungLoaiHH ='"+cl+"'")
                    # mcl=c.fetchone()[0]
                    query_string = "insert into HangHoa (MaHang,TenHang,DVT,DonGia,MaNhomNH,LT) values (?,?,?,?,?,?)"
                    values = (mh,th,dvt,dg,mnnh,lt)
                    c.execute(query_string, values)
                    c.commit()
                    QMessageBox.information(self, "Thông báo", f"Thêm hàng hóa thành công!")
            else:
                QMessageBox.information(self, "Thông báo", f"Vui lòng nhập đủ thông tin!")
            # clear form input
            self.ui.maHang.setText("")
            # self.ui.maHangTD.setText("")
            self.ui.tenHang.setText("")
            self.ui.dvt.setText("")
            self.ui.donGia.setText("")
            self.ui.nhomNH.setCurrentIndex(0)
            self.ui.leadTime.setText("")
            # load new user from db to table view
            AppFunction.displayHH(self)
        except Error as e:
            QMessageBox.information(self, "Thông báo", f"Thêm hàng hóa thất bại!")
    def SuaHH(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        mh=self.ui.maHang.toPlainText()
        # mhtd=self.ui.maHangTD.toPlainText()
        th=self.ui.tenHang.toPlainText()
        dvt=self.ui.dvt.toPlainText()
        dg=self.ui.donGia.toPlainText()
        nnh=self.ui.nhomNH.currentText()
        lt=self.ui.leadTime.toPlainText()
        if len(mh)>0 and len(th)>0 and len(dvt)>0 and len(dg)>0 and nnh!="Nhóm ngành hàng" and len(lt)>0: 
            try:
                c.execute("select MaNhomNH from NhomNH where TenNhomNH ='"+nnh+"'")
                mnnh=c.fetchone()[0]
                # c.execute("select MaChungLoaiHH from ChungLoaiHH where TenChungLoaiHH ='"+cl+"'")
                # mcl=c.fetchone()[0]
                c.execute("update HangHoa set TenHang='{}',DVT='{}',DonGia={},MaNhomNH={},LT={} where MaHang={}".format( th, dvt,dg,mnnh,lt,mh))
                c.commit()
                QMessageBox.information(self, "Thông báo", f"Sửa thông tin hàng hóa thành công!")
                AppFunction.displayHH(self)
            except:
                QMessageBox.information(self, "Thông báo", f"Sửa thông tin hàng hóa thất bại!")
        else:
            QMessageBox.information(self, "Thông báo", f"Vui lòng thử lại!")
        # clear form input
        self.ui.maHang.setText("")
        # self.ui.maHangTD.setText("")
        self.ui.tenHang.setText("")
        self.ui.dvt.setText("")
        self.ui.donGia.setText("")
        self.ui.nhomNH.setCurrentIndex(0)
        self.ui.leadTime.setText("")
    def XoaHH(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        row=self.ui.tbHangHoa.currentRow()
        mh=self.ui.tbHangHoa.item(int(row),0).text()
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Question)
        msgBox.setText(f"Bạn có chắc chắn muốn xóa?")
        msgBox.setWindowTitle("Thông báo")
        msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        result = msgBox.exec_()
        if result == QMessageBox.Ok:
            try:
                c.execute("Delete from HangHoa where MaHang='"+mh+"'")
                c.commit()
                QMessageBox.information(self, "Thông báo", f"Xóa hàng hóa thành công!")
            except:
                QMessageBox.information(self, "Thông báo", f"Xóa hàng hóa thất bại!")
        AppFunction.displayHH(self)
    def NapHH(self):
        row=self.ui.tbHangHoa.currentRow()
        mh=self.ui.tbHangHoa.item(int(row),0).text()
        # mhtd=self.ui.tbHangHoa.item(int(row),1).text()
        th=self.ui.tbHangHoa.item(int(row),1).text()
        dvt=self.ui.tbHangHoa.item(int(row),2).text()
        dg=self.ui.tbHangHoa.item(int(row),3).text()
        nnh=self.ui.tbHangHoa.item(int(row),4).text()
        lt=self.ui.tbHangHoa.item(int(row),5).text()
        self.ui.maHang.setText(mh)
        # self.ui.maHangTD.setText(mhtd)
        self.ui.tenHang.setText(th)
        self.ui.dvt.setText(dvt)
        self.ui.donGia.setText(dg)
        self.ui.nhomNH.setCurrentText(nnh)
        self.ui.leadTime.setText(lt)
    # Bảng cân đối vật tư
    def CDVT(self):
        global dataframe
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        c.execute("with DA_HMDT as (select MaDA,TenHMDT from DA,HMDT where DA.MaHMDT=HMDT.MaHMDT) select TenNhomNH,A010.MaHang,TenHang,SLDat,TenHMDT from NhomNH,A010, DA_HMDT,HangHoa where DA_HMDT.MaDA=A010.MaDA and HangHoa.MaHang=A010.MaHang and HangHoa.MaNhomNH=NhomNH.MaNhomNH")
        dfA010=c.fetchall()
        data =[]
        for i in dfA010:
            i=tuple(i)
            data.append(i)
        names = [ x[0] for x in c.description]
        dfA010 = pd.DataFrame(data, columns=names)
        dfA010["MaHang"]=dfA010['MaHang'].str.strip()
        dfA010 = dfA010.groupby(["TenNhomNH","MaHang", "TenHang","TenHMDT"], as_index=False)["SLDat"].sum()
        dfA010=dfA010.pivot(index=['TenNhomNH','MaHang','TenHang'],columns='TenHMDT',values='SLDat')
        dfA010=dfA010.fillna(0)
        dfA010.columns.name=None
        dfA010=dfA010.reset_index()
        dfA010['TongSL']=dfA010['Bảo trì, XLSC, Cải tạo hạ tầng']+dfA010['Công cụ dụng cụ']+dfA010['Dự án Metro POP, POP +']+dfA010['Dự án Ngầm hóa']+dfA010['Nâng cấp hạ tầng Access']+dfA010['Phát triển hạ tầng']+dfA010['Dự án/ KH phát sinh khác']
        dfDPXLSC=AppFunction.ImportQuotaDPXLSC(self)
        dfTonNCC=AppFunction.ImportTonNCC(self)
        dfTonNCC['Mã hàng']=dfTonNCC['Mã hàng'].astype(str)
        dfDPXLSC['Mã hàng']=dfDPXLSC['Mã hàng'].astype(str)
        c.execute("Select MaHang,sum(SLTonKD) as SLTonMB from B048 group by MaHang")
        dfTonMB=c.fetchall()
        data1 =[]
        for i in dfTonMB:
            i=tuple(i)
            data1.append(i)
        names1 = [ x[0] for x in c.description]
        dfTonMB = pd.DataFrame(data1, columns=names1)
        dfTonMB["MaHang"]=dfTonMB["MaHang"].str.strip()
        # dfDPXLSC["Mã hàng"]=dfDPXLSC["Mã hàng"].str.strip()
        dfTonNCC["Mã hàng"]=dfTonNCC["Mã hàng"].str.strip()
        df=pd.merge(dfA010,dfTonNCC,left_on='MaHang',right_on='Mã hàng',how='left')
        df.drop(columns=['Mã hàng','Tên hàng'],inplace=True)
        df.rename(columns={'Số lượng':'TonNCC'},inplace=True)
        df=pd.merge(df,dfDPXLSC,left_on='MaHang',right_on='Mã hàng',how='left')
        df.drop(columns=['Mã hàng'],inplace=True)
        df.rename(columns={'Số lượng':'DPXLSC'},inplace=True)
        df=pd.merge(df,dfTonMB,on='MaHang',how='left')
        df=df.fillna(0)
        df['SLSauCD']=(df['SLTonMB']+df['TonNCC'])-(df['TongSL']+df['DPXLSC'])
        df['TTMua']=np.where(df['SLSauCD']>0,'Không mua','Mua')
        df['SLTonMB']=np.round(df['SLTonMB'],1)
        df['SLSauCD']=np.round(df['SLSauCD'],1)
        # cách dưới khác oke
        # header = self.ui.tbSoatDatHang.horizontalHeader()
        # header.setSectionResizeMode(QHeaderView.ResizeToContents)

        self.ui.tbSoatDatHang.setRowCount(len(df))
        for row_number, (_,row_data) in enumerate(df.iterrows()):
            for column_number, data in enumerate(row_data):
                self.ui.tbSoatDatHang.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        delegate = AlignRDelegate()
        for i in range(3,len(df.columns)):
            self.ui.tbSoatDatHang.setItemDelegateForColumn(i, delegate)
        dataframe=df
    def Delete(self):
        selected_items = self.ui.tbSoatDatHang.selectedItems()
        if len(selected_items) == 0:
            # Hiển thị thông báo yêu cầu chọn một dòng trước khi xóa
            return

        rows = set()
        for item in selected_items:
            rows.add(item.row())

        for row in sorted(rows, reverse=True):
            self.ui.tbSoatDatHang.removeRow(row)
    def FilterCDVT(self):
        global dataframe
        df=dataframe
        cb=self.ui.muaHangCb.currentText()
        
        df=df[df["TTMua"]==cb]
        self.ui.tbSoatDatHang.setRowCount(len(df))
        for row_number, (_,row_data) in enumerate(df.iterrows()):
            for column_number, data in enumerate(row_data):
                self.ui.tbSoatDatHang.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        delegate = AlignRDelegate()
        for i in range(3,len(df.columns)):
            self.ui.tbSoatDatHang.setItemDelegateForColumn(i, delegate)
        return df
    # def DisplayDH(self):
    #     conn=AppFunction.create_connection(self)
    #     c=conn.cursor()
    #     c.execute("Select MaDH, TenNV,NgayDatDH,TenCN,DienGiai,TenTT,NamTC,MaDA,A010Detail.MaHang,TenHang,SLDat,SLDuyet,SLChuaGiao from A010,NhanVien,CN,TrangThai,HangHoa where A010.MaNVTao=NhanVien.MaNV and A010.MaCN=CN.MaCN and A010.MaTT=TrangThai.MaTT and A010.MaHang=HangHoa.MaHang")
    #     dfA010=c.fetchall()
    #     data =[]
    #     for i in dfA010:
    #         i=tuple(i)
    #         data.append(i)
    #     names = [ x[0] for x in c.description]
    #     dfA010 = pd.DataFrame(data, columns=names)
    #     self.ui.tbSoatDatHangTruoc.setRowCount(len(dfA010))
    #     for row_number, (_,row_data) in enumerate(dfA010.iterrows()):
    #         for column_number, data in enumerate(row_data):
    #             self.ui.tbSoatDatHangTruoc.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
    #     delegate = AlignRDelegate()
    #     for i in range(10,len(dfA010.columns)):
    #         self.ui.tbSoatDatHangTruoc.setItemDelegateForColumn(i, delegate)
    def DataVisual(self,mpl_canvas):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        c.execute("with DAHMDT as(select MaDA,TenHMDT from DA,HMDT where DA.MaHMDT=HMDT.MaHMDT) select * from B048 left outer join DAHMDT on B048.MaDA=DAHMDT.MaDA")
        dfB048=c.fetchall()
        data =[]
        for i in dfB048:
            i=tuple(i)
            data.append(i)
        names = [ x[0] for x in c.description]
        dfB048 = pd.DataFrame(data, columns=names)
        c.execute("select MaHang,DonGia,TenNhomNH from HangHoa,NhomNH where HangHoa.MaNhomNH=NhomNH.MaNhomNH")
        dfHH=c.fetchall()
        data1 =[]
        for i in dfHH:
            i=tuple(i)
            data1.append(i)
        names1 = [ x[0] for x in c.description]
        dfHH = pd.DataFrame(data1, columns=names1)
        c.execute("select MaKho,TenVung,TenVietTat from Kho,CN,Vung where Kho.MaCN=CN.MaCN and CN.MaVung=Vung.MaVung")
        dfKho=c.fetchall()
        data2 =[]
        for i in dfKho:
            i=tuple(i)
            data2.append(i)
        names2 = [ x[0] for x in c.description]
        dfKho = pd.DataFrame(data2, columns=names2)
        dfB048=pd.merge(dfB048,dfHH,on="MaHang",how="inner")
        dfB048=pd.merge(dfB048,dfKho,on="MaKho",how="inner")
        dfB048["Giá trị tồn khả dụng"]=float(0) 
        dfB048["Giá trị tổng tồn"]=float(0)
        dfB048["Giá trị tồn > 90d"]=float(0)
        for i in range (len(dfB048["MaTTH"])):
            if dfB048.loc[i,'MaTTH'] == 'TT029' or dfB048.loc[i,'MaTTH']== 'TT039':
                dfB048.loc[i,"Giá trị tồn khả dụng"]=(dfB048.loc[i,"SLTonKD"]*dfB048.loc[i,"DonGia"])/1000000000
            else: 
                dfB048.loc[i,"Giá trị tồn khả dụng"]=0
        for i in range (len(dfB048["MaTTH"])):
            if dfB048.loc[i,'MaTTH'] == 'TT029' or dfB048.loc[i,'MaTTH']== 'TT039':
                dfB048.loc[i,"Giá trị tổng tồn"]=(dfB048.loc[i,"TongTon"]*dfB048.loc[i,"DonGia"])/1000000000
            else: 
                dfB048.loc[i,"Giá trị tổng tồn"]=0
        for i in range (len(dfB048["MaTTH"])):
            if dfB048.loc[i,'MaTTH'] == 'TT029' or dfB048.loc[i,'MaTTH']== 'TT039':
                dfB048.loc[i, "Giá trị tồn > 90d"] = (dfB048.loc[i, "SLTonHon90"] * dfB048.loc[i, "DonGia"]) / 1000000000
            else: 
                dfB048.loc[i,"Giá trị tồn > 90d"]=0
        dfB048["Giá trị tồn > 90d"] = dfB048["Giá trị tồn > 90d"].astype(float)
        dfB048["Giá trị tổng tồn"] = dfB048["Giá trị tổng tồn"].astype(float)
        dfB048["Giá trị tồn khả dụng"] = dfB048["Giá trị tồn khả dụng"].astype(float)
        dfB048["TenHMDT"].fillna("Chưa gắn HMĐT",inplace=True)
        return dfB048
    def ExportBCCDVT(self):
        # df=AppFunction.FilterCDVT(self)
        df = pd.DataFrame()
        num_rows = self.ui.tbSoatDatHang.rowCount()
        num_cols = self.ui.tbSoatDatHang.columnCount()
        # Lấy tên cột từ TableWidget
        columns = []
        for col in range(num_cols):
            header_item = self.ui.tbSoatDatHang.horizontalHeaderItem(col)
            if header_item is not None:
                columns.append(header_item.text())
            else:
                columns.append(f"Column {col+1}")

        # Tạo DataFrame với các cột đã lấy được
        df = pd.DataFrame(columns=columns)

        # Lặp qua từng hàng của TableWidget
        for row in range(num_rows):
            data = []
            for col in range(num_cols):
                item = self.ui.tbSoatDatHang.item(row, col)
                if item is not None:
                    data.append(item.text())
                else:
                    data.append('')
            # Thêm dữ liệu hàng hiện tại vào DataFrame
            df.loc[row] = data
        # num_rows = self.ui.tbSoatDatHang.rowCount()
        # num_cols = self.ui.tbSoatDatHang.columnCount()
        # for row in range(num_rows):
        #     data1 = []
        #     for col in range(num_cols):
        #         item = self.ui.tbSoatDatHang.item(row, col)
        #         if item is not None:
        #             data1.append(item.text())
        #         else:
        #             data1.append('')
        #     # df = df._append(pd.Series(data1), ignore_index=True)
        #     # Tạo một DataFrame tạm thời từ dữ liệu hàng hiện tại
        #     temp_df = pd.DataFrame([data1], columns=df.columns)
        #     # Sử dụng phương thức concat để nối DataFrame tạm thời vào DataFrame chính
        #     df = pd.concat([df, temp_df], ignore_index=True)
        df['SL sau cân đối'] = df['SL sau cân đối'].astype(float)
        df['SL sau cân đối']=df['SL sau cân đối']*(-1)
        df=df[['Nhóm NH','Mã hàng','Tên hàng','SL sau cân đối']]
        data=df.values.tolist()
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        df.rename(columns={'Mã hàng':'MaHang'},inplace=True)
        c.execute("Select MaHang,DonGia from HangHoa")
        dfHH=c.fetchall()
        data1 =[]
        for i in dfHH:
            i=tuple(i)
            data1.append(i)
        names = [ x[0] for x in c.description]
        dfHH = pd.DataFrame(data1, columns=names)
        dfHH['MaHang']=dfHH['MaHang'].str.strip()
        df=pd.merge(df,dfHH,on="MaHang",how='inner')
        # df['DonGia'] = df['DonGia'].astype(float)
        df["Giá trị"]=df["SL sau cân đối"]*df["DonGia"]
        col=['Nhóm hàng','Mã hàng','Tên hàng','Số lượng đặt']
        pdf = PDFDatHang()
        pdf.add_page()
        pdf.alias_nb_pages()
        pdf.set_auto_page_break(False)
        pdf.set_widths([50,30,70,40])
        pdf.set_font('DejaVu',size=10)
        pdf.set_x(10)
        pdf.multi_cell(0, 5,'Công ty Cổ phần Viễn thông FPT \nTrung tâm Phát triển và Quản lý hạ tầng MB \nĐịa chỉ: 48 Vạn Bảo, Ngọc Khánh, Ba Đình', border="B", align='L')
        pdf.ln(5) 
        pdf.set_font('DejaVu', 'B', 15)
        pdf.multi_cell(0, 7, 'BÁO CÁO SỐ LƯỢNG ĐẶT HÀNG', border=0, align='C')
        pdf.ln(3)
        pdf.set_font('DejaVu', '', 10)
        tz_VN = pytz.timezone('Asia/Ho_Chi_Minh')
        datetime_VN = datetime.now(tz_VN)
        now_vn = datetime_VN.strftime("%d/%m/%Y")
        pdf.multi_cell(0, 7, f'Ngày: {now_vn}', border=0, align='R')
        pdf.ln(5)
        pdf.tieude(col) 
        for i in range(len(data)):
            if pdf.get_y() + 30 > pdf.h:
                pdf.add_page()
                pdf.tieude(col)
            pdf.row([data[i]])
        sum1=df["Giá trị"].sum()
        tong=["Tổng giá trị",sum1]
        pdf.sum(tong)
        filename = filedialog.asksaveasfilename(defaultextension='.pdf')
        if filename:
            pdf.output(filename, 'F')
    def ImportQuotaDPXLSC(self):
        filename=filedialog.askopenfilename()
        if filename:
            df=pd.read_excel(filename)
        df['Mã hàng']=df['Mã hàng'].astype(str)
        df = df.groupby(["Mã hàng"], as_index=False)["Số lượng"].sum()
        return df
    def ImportTonNCC(self):
        filename=filedialog.askopenfilename()
        if filename:
            df=pd.read_excel(filename)
        return df
    def ChangePW(self):
        ttk=self.ui.tenTKDoiMK.text()
        pwnew=self.ui.mkDoiMK.toPlainText()
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        try:
            c.execute("update TK set MK='"+pwnew+"' where TenTK='"+ttk+"'")
            c.commit()
            QMessageBox.information(self, "Thông báo", f"Đối mật khẩu thành công!")
        except:
            QMessageBox.information(self, "Thông báo", f"Đối mật khẩu thất bại!")
    # Cập nhật data
    def un_pad(plain_text):
        last_character = plain_text[len(plain_text) - 1:]
        bytes_to_remove = ord(last_character)
        return plain_text[:-bytes_to_remove]
    # Hàm giải mã
    def aes_decrypt(plan_text):
        cipher = AES.new(AppFunction.AES_KEY.encode('utf8'),
                        AES.MODE_CBC,
                        AppFunction.AES_IV.encode('utf8'))
        plain_text = cipher.decrypt(base64.b64decode(plan_text)).decode('utf8')
        return AppFunction.un_pad(plain_text)
    # Hàm mã hóa
    def aes_encrypt(plan_text):
        msg = pad(str(plan_text).encode('utf-8'), AES.block_size)
        cipher = AES.new(AppFunction.AES_KEY.encode('utf8'),
                        AES.MODE_CBC,
                        AppFunction.AES_IV.encode('utf8'))
        cipher_text = base64.b64encode(cipher.encrypt(msg)).decode('utf-8')
        return cipher_text
    def handle_api_B048(self):
        # date_temp=self.ui.toDate.date()
        # today=date_temp.toPyDate()
        tz_VN = pytz.timezone('Asia/Ho_Chi_Minh')
        datetime_VN = datetime.now(tz_VN)
        now_vn = datetime_VN.strftime("%d%m%y")
        today=now_vn
        content = """
        "PageSize": 3000,
        "LIST_ITEM_CODE": null,
        "FromRecord": 1,
        "VIEW_TYPE": 0,
        "V_REPORT": 48,
        "TO_DATE": "{today}",
        "LIST_LOCATION_ID": null,
        "LIST_STOCK_ID": null,
        "LIST_ZONE_ID": "29,30,31,39"
        """
        encrypted_content = AppFunction.aes_encrypt('{' + content + '}')
        payload = {
            "Header": {
                "ApiVersion": "1.0",
                "MessageId": AppFunction.AES_IV,
                "RequestFrom": AppFunction.SCM_USER,
                "RequestTo": "FTEL.SCM",
                "RequestTime": "2023-10-27",
                "ParentBusiness": "019",
                "ChildBusiness": "007"
            },
            "Body": {
                "Content": encrypted_content
            },
            "Footer": {}
        }
        headers = {
            'Content-Type': 'application/json'
        }
        response = requests.request("POST",
                                    AppFunction.SCM_API,
                                    headers=headers,
                                    data=json.dumps(payload))
        response = json.loads(response.text)
        footer_sys_code = response.get('Footer').get('SysCode')
        if footer_sys_code != '0':
            footer_sys_message = response.get('Footer').get('SysMessage')
            return {'is_error': True, 'message': footer_sys_message}
        else:
            body_content = response.get('Body').get('Content')
            body_content = AppFunction.aes_decrypt(body_content)
            body_content = json.loads(body_content)
            body_content_code = body_content.get('Code')
            body_content_message = body_content.get('Message')
            total = body_content.get('Total')
            PageIndex= body_content.get('PageIndex')
            total_pages = total // 3000 + 1
            all_data = []
            for PageIndex in range(1, total_pages + 1):
                record=(PageIndex-1)*int(3000)+1
                content = """
                    "PageSize": 3000,
                    "LIST_ITEM_CODE": null,
                    "FromRecord": 1,
                    "VIEW_TYPE": 0,
                    "V_REPORT": 48,
                    "TO_DATE": "{today}",
                    "LIST_LOCATION_ID": null,
                    "LIST_STOCK_ID": null,
                    "LIST_ZONE_ID": "29,30,31,39"
                        """
                content = content.replace('"FromRecord": 1', f'"FromRecord": {record}')
                encrypted_content = AppFunction.aes_encrypt('{' + content + '}')
                payload["Body"]["Content"] = encrypted_content
                response = requests.request("POST",
                                    AppFunction.SCM_API,
                                    headers=headers,
                                    data=json.dumps(payload))
                response = json.loads(response.text)
                footer_sys_code = response.get('Footer').get('SysCode')
                if footer_sys_code != '0':
                    footer_sys_message = response.get('Footer').get('SysMessage')
                    return {'is_error': True, 'message': footer_sys_message}
                else:
                    body_content = response.get('Body').get('Content')
                    body_content = AppFunction.aes_decrypt(body_content)
                    body_content = json.loads(body_content)
                    body_content_code = body_content.get('Code')
                    body_content_message = body_content.get('Message')
                    all_data.append(body_content)
        return all_data
    def handle_api_A010(self):
        # date_temp=self.ui.fromDate.date()
        # fromdate=date_temp.toPyDate()
        # date_temp1=self.ui.toDate.date()
        # todate=date_temp1.toPyDate()
        # namtc=self.ui.namTC.toPlainText()
        tz_VN = pytz.timezone('Asia/Ho_Chi_Minh')
        datetime_VN = datetime.now(tz_VN)
        now_vn = datetime_VN.strftime("%d%m%y")
        namtc = datetime_VN.strftime("%Y")
        fromdate='01/01/{namtc}'
        todate=now_vn
        content = """
        "PageIndex": 1,
        "PageSize": 3000,
        "List_DEPARTMENT_ID": null,
        "List_REGIONS_ID": "33,34,35",
        "List_PROJECT_ID": null,
        "List_ORDER_TYPE_ID":null,
        "List_STATE_ID": 627,
        "List_STATUS": 4,
        "LIST_ITEM_CODE": null,
        "List_OM_HEADER_ID": null,
        "List_PERIOD_YEAR": {namtc},
        "DATE_TYPE": 0,
        "List_INVESTMENT_ID": null,
        "REQUEST_DATE_FROM": "{fromdate}",
        "REQUEST_DATE_TO": "{todate}"
        """
        encrypted_content = AppFunction.aes_encrypt('{' + content + '}')
        payload = {
            "Header": {
                "ApiVersion": "1.0",
                "MessageId": AppFunction.AES_IV,
                "RequestFrom": AppFunction.SCM_USER,
                "RequestTo": "FTEL.SCM",
                "RequestTime": "2023-08-16",
                "ParentBusiness": "019",
                "ChildBusiness": "004"
            },
            "Body": {
                "Content": encrypted_content
            },
            "Footer": {}
        }
        headers = {
            'Content-Type': 'application/json'
        }
        response = requests.request("POST",
                                    AppFunction.SCM_API,
                                    headers=headers,
                                    data=json.dumps(payload))
        response = json.loads(response.text)
        footer_sys_code = response.get('Footer').get('SysCode')
        if footer_sys_code != '0':
            footer_sys_message = response.get('Footer').get('SysMessage')
            return {'is_error': True, 'message': footer_sys_message}
        else:
            body_content = response.get('Body').get('Content')
            body_content = AppFunction.aes_decrypt(body_content)
            body_content = json.loads(body_content)
            body_content_code = body_content.get('Code')
            body_content_message = body_content.get('Message')
            total = body_content.get('Total')
            total_pages = total // 3000 + 1
            all_data = []
            for page in range(1, total_pages + 1):
                content = """
                    "PageIndex": 1,
                    "PageSize": 3000,
                    "List_DEPARTMENT_ID": null,
                    "List_REGIONS_ID": "33,34,35",
                    "List_PROJECT_ID": null,
                    "List_ORDER_TYPE_ID":null,
                    "List_STATE_ID": 627,
                    "List_STATUS": 4,
                    "LIST_ITEM_CODE": null,
                    "List_OM_HEADER_ID": null,
                    "List_PERIOD_YEAR": null,
                    "DATE_TYPE": 0,
                    "List_INVESTMENT_ID": null,
                    "REQUEST_DATE_FROM": "{fromdate}",
                    "REQUEST_DATE_TO": "{todate}"
                        """
                content = content.replace('"PageIndex": 1', f'"PageIndex": {page}')
                encrypted_content = AppFunction.aes_encrypt('{' + content + '}')
                payload["Body"]["Content"] = encrypted_content
                response = requests.request("POST",
                                    AppFunction.SCM_API,
                                    headers=headers,
                                    data=json.dumps(payload))
                response = json.loads(response.text)
                footer_sys_code = response.get('Footer').get('SysCode')
                if footer_sys_code != '0':
                    footer_sys_message = response.get('Footer').get('SysMessage')
                    return {'is_error': True, 'message': footer_sys_message}
                else:
                    body_content = response.get('Body').get('Content')
                    body_content = AppFunction.aes_decrypt(body_content)
                    body_content = json.loads(body_content)
                    body_content_code = body_content.get('Code')
                    body_content_message = body_content.get('Message')
                    all_data.append(body_content)
        return all_data
    def handle_api_A040(self):
        # date_temp=self.ui.fromDate.date()
        # fromdate=date_temp.toPyDate()
        # date_temp1=self.ui.toDate.date()
        # todate=date_temp1.toPyDate()
        # namtc=self.ui.namTC.toPlainText()
        tz_VN = pytz.timezone('Asia/Ho_Chi_Minh')
        datetime_VN = datetime.now(tz_VN)
        now_vn = datetime_VN.strftime("%d%m%y")
        namtc = datetime_VN.strftime("%Y")
        fromdate='01/01/{namtc}'
        todate=now_vn
        content = """
        "PageIndex": 1,
        "PageSize": 3000,
        "ProcessId": 4,
        "List_INVESTMENT_ID": null,
        "List_PROJECT_ID": null,
        "List_PERIOD_YEAR": "{namtc}",
        "List_DEPARTMENT_ID": null,
        "List_REGIONS_ID": "33,34,35",
        "List_BRANCHES_ID": null,
        "List_STATE_ID": "88,89",
        "LIST_OM_REQUEST_CODE": null,
        "LIST_ITEM_CODE": null,
        "LIST_ORDER_TYPE_ID": null,
        "LIST_ORDER_TYPE_ID2": null,
        "PAR_DELIVERY_CODE": null,
        "REQUEST_DATE_FROM": "{fromdate}",
        "REQUEST_DATE_TO": "{todate}",
        "FromRecord": 0,
        "IS_EXPORT_EXCEL_DETAIL": 1
        """
        encrypted_content = AppFunction.aes_encrypt('{' + content + '}')
        payload = {
            "Header": {
                "ApiVersion": "1.0",
                "MessageId": AppFunction.AES_IV,
                "RequestFrom": AppFunction.SCM_USER,
                "RequestTo": "FTEL.SCM",
                "RequestTime": "2021-07-28T17:32:35+07:00",
                "ParentBusiness": "019",
                "ChildBusiness": "006"
            },
            "Body": {
                "Content": encrypted_content
            },
            "Footer": {}
        }
        headers = {
            'Content-Type': 'application/json'
        }
        response = requests.request("POST",
                                    AppFunction.SCM_API,
                                    headers=headers,
                                    data=json.dumps(payload))
        response = json.loads(response.text)
        footer_sys_code = response.get('Footer').get('SysCode')
        if footer_sys_code != '0':
            footer_sys_message = response.get('Footer').get('SysMessage')
            return {'is_error': True, 'message': footer_sys_message}
        else:
            body_content = response.get('Body').get('Content')
            body_content = AppFunction.aes_decrypt(body_content)
            body_content = json.loads(body_content)
            body_content_code = body_content.get('Code')
            body_content_message = body_content.get('Message')
            total = body_content.get('Total')
            total_pages = total // 3000 + 1
            all_data = []
            for page in range(1, total_pages + 1):
                content = """
                    "PageIndex": 1,
                    "PageSize": 3000,
                    "ProcessId": 4,
                    "List_INVESTMENT_ID": null,
                    "List_PROJECT_ID": null,
                    "List_PERIOD_YEAR": "{namtc}",
                    "List_DEPARTMENT_ID": null,
                    "List_REGIONS_ID": "33,34,35",
                    "List_BRANCHES_ID": null,
                    "List_STATE_ID": "88,89",
                    "LIST_OM_REQUEST_CODE": null,
                    "LIST_ITEM_CODE": null,
                    "LIST_ORDER_TYPE_ID": null,
                    "LIST_ORDER_TYPE_ID2": null,
                    "PAR_DELIVERY_CODE": null,
                    "REQUEST_DATE_FROM": "{fromdate}",
                    "REQUEST_DATE_TO": "{todate}",
                    "FromRecord": 0,
                    "IS_EXPORT_EXCEL_DETAIL": 1
                        """
                content = content.replace('"PageIndex": 1', f'"PageIndex": {page}')
                encrypted_content = AppFunction.aes_encrypt('{' + content + '}')
                payload["Body"]["Content"] = encrypted_content
                response = requests.request("POST",
                                    AppFunction.SCM_API,
                                    headers=headers,
                                    data=json.dumps(payload))
                response = json.loads(response.text)
                footer_sys_code = response.get('Footer').get('SysCode')
                if footer_sys_code != '0':
                    footer_sys_message = response.get('Footer').get('SysMessage')
                    return {'is_error': True, 'message': footer_sys_message}
                else:
                    body_content = response.get('Body').get('Content')
                    body_content = AppFunction.aes_decrypt(body_content)
                    body_content = json.loads(body_content)
                    body_content_code = body_content.get('Code')
                    body_content_message = body_content.get('Message')
                    all_data.append(body_content)
        return all_data
    def handle_api_B022(self):
        tz_VN = pytz.timezone('Asia/Ho_Chi_Minh')
        datetime_VN = datetime.now(tz_VN)
        now_vn = datetime_VN.strftime("%d%m%y")
        namtc = datetime_VN.strftime("%Y")
        fromdate='01/01/{namtc}'
        todate=now_vn
        content = """
        "PageIndex": 1,
        "PageSize": 3000,
        "LIST_STOCK_ID": null,
        "LIST_ITEM_CODE": null,
        "FROM_DATE": "{fromdate}",
        "TO_DATE": "{todate}",
        "LIST_EX_STOCK_ID": null,
        "LIST_ZONE_ID": null,
        "LIST_GROUP_ID": null,
        "LIST_PROJECT_ID": null,
        "LIST_PLAN_ID": null,
        "LIST_INCIDIENT_ID": null,
        "LIST_STOCK_TYPE_ID": null,
        "LIST_DEPARTMENT_ID": null,
        "LIST_EMPLOYEE_ID": null,
        "P_TYPE_VIEW": 3,
        "MAC_TK": null,
        "SERI_TK": null
        """
        encrypted_content = AppFunction.aes_encrypt('{' + content + '}')
        payload = {
            "Header": {
                "ApiVersion": "1.0",  
                "MessageId": "bscKHn8REOJ2aikS",  
                "RequestFrom": "INF.MAPNET",  
                "RequestTo": "FTEL.SCM",  
                "RequestTime": "2021-07-28T17:32:35+07:00",  
                "ParentBusiness": "019",  
                "ChildBusiness": "011"  
            },
            "Body": {
                "Content": encrypted_content
            },
            "Footer": {}
        }
        headers = {
            'Content-Type': 'application/json'
        }
        response = requests.request("POST",
                                    AppFunction.SCM_API,
                                    headers=headers,
                                    data=json.dumps(payload))
        response = json.loads(response.text)
        footer_sys_code = response.get('Footer').get('SysCode')
        if footer_sys_code != '0':
            footer_sys_message = response.get('Footer').get('SysMessage')
            return {'is_error': True, 'message': footer_sys_message}
        else:
            body_content = response.get('Body').get('Content')
            body_content = AppFunction.aes_decrypt(body_content)
            body_content = json.loads(body_content)
            body_content_code = body_content.get('Code')
            body_content_message = body_content.get('Message')
            total = body_content.get('Total')
            total_pages = total // 3000 + 1
            all_data = []
            for page in range(1, total_pages + 1):
                content = """
                    "PageIndex": 1,
                    "PageSize": 3000,
                    "LIST_STOCK_ID": null,
                    "LIST_ITEM_CODE": null,
                    "FROM_DATE": "{fromdate}",
                    "TO_DATE": "{todate}",
                    "LIST_EX_STOCK_ID": null,
                    "LIST_ZONE_ID": null,
                    "LIST_GROUP_ID": null,
                    "LIST_PROJECT_ID": null,
                    "LIST_PLAN_ID": null,
                    "LIST_INCIDIENT_ID": null,
                    "LIST_STOCK_TYPE_ID": null,
                    "LIST_DEPARTMENT_ID": null,
                    "LIST_EMPLOYEE_ID": null,
                    "P_TYPE_VIEW": 3,
                    "MAC_TK": null,
                    "SERI_TK": null
                        """
                content = content.replace('"PageIndex": 1', f'"PageIndex": {page}')
                encrypted_content = AppFunction.aes_encrypt('{' + content + '}')
                payload["Body"]["Content"] = encrypted_content
                response = requests.request("POST",
                                    AppFunction.SCM_API,
                                    headers=headers,
                                    data=json.dumps(payload))
                response = json.loads(response.text)
                footer_sys_code = response.get('Footer').get('SysCode')
                if footer_sys_code != '0':
                    footer_sys_message = response.get('Footer').get('SysMessage')
                    return {'is_error': True, 'message': footer_sys_message}
                else:
                    body_content = response.get('Body').get('Content')
                    body_content = AppFunction.aes_decrypt(body_content)
                    body_content = json.loads(body_content)
                    body_content_code = body_content.get('Code')
                    body_content_message = body_content.get('Message')
                    all_data.append(body_content)
        return all_data
    def processing_data(self,df):
        df=list(df.values())[0]
        df = pd.json_normalize(df)
        return df
    def capNhatB048(self):
        try:
            df=AppFunction.handle_api_B048(self)
            tz_VN = pytz.timezone('Asia/Ho_Chi_Minh')
            datetime_VN = datetime.now(tz_VN)
            now_vn = datetime_VN.strftime("%d/%m/%Y")
            conn=AppFunction.create_connection(self)
            c=conn.cursor()
            for i,row in df.iterrows():
                mk=row[0]
                mtt=row[1]
                mda=row[2]
                mh=row[3]
                slkd=row[4]
                tongton=row[5]
                tonhon90=row[6]
                mab048='B048'+now_vn
                if pd.isna(mda):
                    query_string = f"insert into B048 (MaTTH,MaHang,SLTonHon90,SLTonKD,TongTon,MaKho,MaB048) values ('{mtt}','{mh}','{tonhon90}','{slkd}','{tongton}','{mk}','{mab048}')"
                else:
                    query_string = f"insert into B048 (MaTTH,MaDA,MaHang,SLTonHon90,SLTonKD,TongTon,MaKho,MaB048) values ('{mtt}','{mda}','{mh}','{tonhon90}','{slkd}','{tongton}','{mk}','{mab048}')"
                c.execute(query_string)
            c.commit()
            QMessageBox.information(self, "Thông báo", f"Cập nhật dữ liệu thành công!")
        except:
            QMessageBox.information(self, "Thông báo", f"Cập nhật dữ liệu thất bại. Vui lòng thử lại!")
            self.message.emit("Cập nhật dữ liệu thành công")
            
        #     tz_VN = pytz.timezone('Asia/Ho_Chi_Minh')
        #     datetime_VN = datetime.now(tz_VN)
        #     now_vn = datetime_VN.strftime("%d/%m/%Y")
        #     message=f"Cập nhật dữ liệu B048 ngày {now_vn} thành công!"
        #     second_window = self.NhanVienWindow(message)
    def CapNhatDataO(self):
        QMessageBox.information(self, "Thông báo", f"Cập nhật dữ liệu thất bại!")
    def capNhatA040(self):
        df=AppFunction.handle_api_A040(self)
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        for i,row in df.iterrows():
            mycgh=row[0]
            nv=row[1]
            nyc=row[2]
            kn=row[3]
            dgiai=row[4]
            hmdt=row[5]
            ntc=row[6]
            tt=row[7]
            dh=row[8]
            mh=row[9]
            slyc=row[10]
            if pd.isna(dgiai):
                query_string = f"insert into A040 (MaYCGH,MaHang,MaTT,MaHMDT,MaDHNoiDVGoc,MaNguoiTao,NgayYC,SLYC,NamTC,MaKhoNhap) values ('{mycgh}','{mh}','{tt}','{hmdt}','{dh}','{nv}','{nyc}','{slyc}','{ntc}','{kn}')"
            else:
                query_string = f"insert into A040 (MaYCGH,MaHang,MaTT,MaHMDT,MaDHNoiDVGoc,MaNguoiTao,NgayYC,SLYC,NamTC,DienGiai,MaKhoNhap) values ('{mycgh}','{mh}','{tt}','{hmdt}','{dh}','{nv}','{nyc}','{slyc}','{ntc}',N'{dgiai}','{kn}')"
            c.execute(query_string)
        c.commit()
    def capNhatA010(self):
        df=AppFunction.handle_api_A010(self)
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        for i,row in df.iterrows():
            mdh=row[0]
            mnvt=row[1]
            mldh=row[2]
            nd=row[3]
            mtt=row[4]
            mda=row[5]
            mh=row[6]
            ntc=row[7]
            mcn=row[8]
            sld=row[9]
            sldd=row[10]
            slcg=row[11]
            mttmh=row[12]
            dgiai=row[13]
            if pd.isna(dgiai):
                query_string = f"insert into A010 (MaDH,MaNVTao, MaLoaiDH, NgayDatDH, MaTT, MaDA, MaHang, NamTC, MaCN, SLDuyet,SLDat , SLChuaGiao, MaTTMuaHang) values ('{mdh}','{mnvt}','{mldh}','{nd}','{mtt}','{mda}','{mh}','{ntc}','{mcn}','{sld}','{sldd}','{slcg}','{mttmh}')"
            else:
                query_string = f"insert into A010 (MaDH,MaNVTao, MaLoaiDH, NgayDatDH,DienGiai, MaTT, MaDA, MaHang, NamTC, MaCN, SLDuyet,SLDat, TgDKGH , SLChuaGiao, MaTTMuaHang) values ('{mdh}','{mnvt}','{mldh}','{nd}','{dgiai}','{mtt}','{mda}','{mh}','{ntc}','{mcn}','{sld}','{sldd}','{slcg}','{mttmh}')"
            c.execute(query_string)
        c.commit()
    def capNhatB022(self):
        df=AppFunction.handle_api_B022(self)
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        for i,row in df.iterrows():
            dng=row[0]
            vung=row[1]
            cn=row[2]
            nttk=row[3]
            mh=row[4]
            th=row[5]
            dvt=row[6]
            slg=row[7]
            dgia=row[8]
        query_string = f"insert into LichSuGD (SoCT,MaHang, MaCN, NgayXK,SLXuat) values ('{dng}','{mh}','{cn}','{nttk}','{slg}')"
        c.execute(query_string)
        c.commit()
    def CapNhatData(self):
        try:
            AppFunction.capNhatA010(self)
            AppFunction.capNhatA040(self)
            AppFunction.capNhatB048(self)
            AppFunction.capNhatB022(self)
            QMessageBox.information(self, "Thông báo", f"Cập nhật dữ liệu thành công!")
        except Error:
            QMessageBox.information(self, "Thông báo", f"Cập nhật dữ liệu thất bại!")
    ######################################## Chức năng chính
    # create 3 grades A,B,C based on the running percentage (A-60%,B-25%,C-15%)
    def ABC_segmentation(self,RunPerc):
        a=self.ui.nhomA.toPlainText()
        b=self.ui.nhomB.toPlainText()
        c=self.ui.nhomC.toPlainText()
        if (a!="") and (b!=""):
            b=float(a)+float(b)
            c=float(100)-float(b)
        self.ui.nhomC.setText(str(c))
        if a == '':
            return 0.0  # or any other default value you want to assign
        else:
            a=float(a)/100
        if b == '':
            return 0.0  # or any other default value you want to assign
        else:
            b=float(b)/100
        if c == '':
            return 0.0  # or any other default value you want to assign
        else:
            c=float(c)/100
        # if (a>1)or (b>1) or (c>1):
        #     msgBox = QMessageBox()
        #     msgBox.setIcon(QMessageBox.Question)
        #     msgBox.setText(f"Đầu vào không hợp lệ!")
        #     msgBox.setWindowTitle("Thông báo")
        #     msgBox.setStandardButtons(QMessageBox.Ok)
        #     result = msgBox.exec_()
        #     if result == QMessageBox.Ok:
        if RunPerc > 0 and RunPerc < a:
            return 'A'
        elif RunPerc >=a and RunPerc < b:
            return 'B'
        elif RunPerc >=b:
            return 'C'
    def ABCAnalysis(self):
        # df=pd.read_excel(r'D:\IMIS\Dashboard\abc.xlsx')
        global dfABC
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        c.execute("select LichSuGD.MaHang,TenHang,DonGia,sum(SLXuat) as NhuCau from LichSuGD,HangHoa where LichSuGD.MaHang=HangHoa.MaHang group by LichSuGD.MaHang,TenHang,DonGia")
        df=c.fetchall()
        data =[]
        for i in df:
            i=tuple(i)
            data.append(i)
        names = [ x[0] for x in c.description]
        df = pd.DataFrame(data, columns=names)
        df['Giá trị'] = df['NhuCau']*df['DonGia']
        # order by AddCost
        df = df.sort_values(by=['Giá trị'], ascending=False)
        # create the column of the running CumCost of the cumulative cost per SKU
        df['RunCumCost'] = df['Giá trị'].cumsum()
        # create the column of the total sum
        df['TotSum'] = df['Giá trị'].sum()
        # create the column of the running percentage 
        df['RunPerc'] = df['RunCumCost'] / df['TotSum']
        # create the column of the class
        df['Class'] = df['RunPerc'].apply(lambda x: AppFunction.ABC_segmentation(self, x))
        df["RunPerc"] = df["RunPerc"].apply(lambda x: x * 100)
        df["RunPerc"] = df["RunPerc"].round(1)
        df1=df[["MaHang","TenHang","DonGia","Giá trị","RunPerc","Class","NhuCau"]]
        dfABC=df1
        df=df[["MaHang","TenHang","DonGia","Giá trị","RunPerc","Class"]]
        self.ui.phanTichABC.setRowCount(len(df))
        for row_number, (_,row_data) in enumerate(df.iterrows()):
            for column_number, data in enumerate(row_data):
                item = QtWidgets.QTableWidgetItem(str(data))
                self.ui.phanTichABC.setItem(row_number, column_number, item)
        self.ui.phanTichABC.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
        self.ui.phanTichABC.setColumnWidth(0, 150)
        self.ui.phanTichABC.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
        self.ui.phanTichABC.setColumnWidth(1, 450)
        self.ui.phanTichABC.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
        self.ui.phanTichABC.setColumnWidth(2, 150)
        self.ui.phanTichABC.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
        self.ui.phanTichABC.setColumnWidth(3, 225)
        self.ui.phanTichABC.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.Fixed)
        self.ui.phanTichABC.setColumnWidth(4, 180)
        self.ui.phanTichABC.horizontalHeader().setSectionResizeMode(5, QtWidgets.QHeaderView.Fixed)
        self.ui.phanTichABC.setColumnWidth(5, 180)
        delegate = AlignRDelegate()
        self.ui.phanTichABC.setItemDelegateForColumn(3, delegate)
        self.ui.phanTichABC.setItemDelegateForColumn(4, delegate)
        self.ui.phanTichABC.setItemDelegateForColumn(5, delegate)
    def TQ(self):
        global dfABC
        def ss_button_clicked():
            df = dfABC
            # total SKUs for each class
            class_counts = df.Class.value_counts().to_frame()
            # print(class_counts)
            class_counts.columns = ['Count']
            class_counts=class_counts.sort_values('Count',ascending=True)
            # print(class_counts)
            cost_a = df[df.Class == 'A']['Giá trị'].sum()
            cost_b = df[df.Class == 'B']['Giá trị'].sum()
            cost_c = df[df.Class == 'C']['Giá trị'].sum()
            cost_a1 = df[df.Class == 'A']['NhuCau'].sum()
            cost_b1 = df[df.Class == 'B']['NhuCau'].sum()
            cost_c1 = df[df.Class == 'C']['NhuCau'].sum()
            percent_a = cost_a / df['Giá trị'].sum()
            percent_b = cost_b / df['Giá trị'].sum()
            percent_c = cost_c / df['Giá trị'].sum()
            percent_a1 = cost_a1 / df['NhuCau'].sum()
            percent_b1 = cost_b1 / df['NhuCau'].sum()
            percent_c1 = cost_c1 / df['NhuCau'].sum()
            data = {'Class': ['A', 'B', 'C'],
                    'Count': class_counts['Count'],
                    'Cost': [cost_a, cost_b, cost_c],
                    'Percent of Total Cost': [percent_a, percent_b, percent_c],
                    'Percent of Total Quantity': [percent_b1, percent_c1, percent_a1]}
            df_summary = pd.DataFrame(data)
            df_summary["Percent of Total Cost"] = df_summary["Percent of Total Cost"].apply(lambda x: x * 100)
            df_summary["Percent of Total Cost"] = df_summary["Percent of Total Cost"].round(1)
            df_summary["Percent of Total Quantity"] = df_summary["Percent of Total Quantity"].apply(lambda x: x * 100)
            df_summary["Percent of Total Quantity"] = df_summary["Percent of Total Quantity"].round(1)
            # print(df_summary)
            self.ui.summary.setRowCount(len(df_summary))
            for row_number, (_,row_data) in enumerate(df_summary.iterrows()):
                for column_number, data in enumerate(row_data):
                    item = QtWidgets.QTableWidgetItem(str(data))
                    self.ui.summary.setItem(row_number, column_number, item)
            self.ui.summary.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
            self.ui.summary.setColumnWidth(0, 150)
            self.ui.summary.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
            self.ui.summary.setColumnWidth(1, 300)
            self.ui.summary.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
            self.ui.summary.setColumnWidth(2, 250)
            self.ui.summary.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
            self.ui.summary.setColumnWidth(3, 250)
            self.ui.summary.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.Fixed)
            self.ui.summary.setColumnWidth(4, 300)
            delegate = AlignRDelegate()
            self.ui.summary.setItemDelegateForColumn(0, delegate)
            self.ui.summary.setItemDelegateForColumn(1, delegate)
            self.ui.summary.setItemDelegateForColumn(2, delegate)
            self.ui.summary.setItemDelegateForColumn(3, delegate)
            self.ui.summary.setItemDelegateForColumn(4, delegate)
        self.ui.tquan.clicked.connect(ss_button_clicked)
    def SS(self):
        # df=pd.read_excel(r'D:\IMIS\Dashboard\ss.xlsx')
        global dfSS
        global dfABC
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        c.execute("select LichSuGD.MaHang,TenHang,NgayXK, sum(SLXuat) as NhuCau from LichSuGD,HangHoa where LichSuGD.MaHang=HangHoa.MaHang group by LichSuGD.MaHang,TenHang,DonGia,NgayXK")
        df1=c.fetchall()
        data =[]
        for i in df1:
            i=tuple(i)
            data.append(i)
        names = [ x[0] for x in c.description]
        df1 = pd.DataFrame(data, columns=names)
        a = df1.groupby(['MaHang']).NhuCau.agg(['mean', lambda x: x.std(ddof=0)])
        a.columns=['mean','std']
        a['std']=a['std'].apply(lambda x: round(x, 2))
        a['mean']=a['mean'].apply(lambda x: round(x, 2))
        a=a.reset_index()
        a['std']=np.where(a['std']==0,1,a['std'])
        dfABC_SS= dfABC
        df=pd.merge(a,dfABC_SS,on='MaHang',how='inner')
        # setup service level
        a=self.ui.sla.toPlainText()
        b=self.ui.slb.toPlainText()
        d=self.ui.slc.toPlainText()
        if a == '':
            return 0.0  # or any other default value you want to assign
        # elif float(a)>100:
        #     QMessageBox.information(self, "Thông báo", f"Đầu vào không hợp lệ!")
        else:
            a=float(a)/100
        if b == '':
            return 0.0  # or any other default value you want to assign
        # elif float(b)>100:
        #     QMessageBox.information(self, "Thông báo", f"Đầu vào không hợp lệ!")
        else:
            b=float(b)/100
        if d == '':
            return 0.0  # or any other default value you want to assign
        # elif float(d)>100:
        #     QMessageBox.information(self, "Thông báo", f"Đầu vào không hợp lệ!")
        else:
            d=float(d)/100
        if (a>1) or (b>1) or (d>1):
            QMessageBox.information(self, "Thông báo", f"Đầu vào không hợp lệ!")
        else:
            df['ServiceLevel']=np.where(df['Class']=='A',a,np.where(df['Class']=='B',b,d))
            df['R']= np.round(norm.ppf(df['ServiceLevel'],df['mean'],df['std']),decimals=2)
            df['z']=(df['R']-df['mean'])/df['std']
            df['SS']=df['z']*df['std']
            c.execute("select MaHang,TongTon from B048")
            dfB048=c.fetchall()
            data1 =[]
            for i in dfB048:
                i=tuple(i)
                data1.append(i)
            names1 = [ x[0] for x in c.description]
            dfB048 = pd.DataFrame(data1, columns=names1)
            c.execute("select MaHang,SLYC from A040Detail")
            dfYCGH=c.fetchall()
            data2 =[]
            for i in dfYCGH:
                i=tuple(i)
                data2.append(i)
            names2 = [ x[0] for x in c.description]
            dfYCGH = pd.DataFrame(data2, columns=names2)
            dfYCGH=dfYCGH.groupby(['MaHang'])['SLYC'].sum()
            dfYCGH=dfYCGH.reset_index()
            # print(dfYCGH)
            dfB048=dfB048.groupby(['MaHang'])['TongTon'].sum()
            dfB048=dfB048.reset_index()
            # print(dfB048)
            dfB048=dfB048.merge(dfYCGH,on='MaHang',how='left')
            dfB048=dfB048.fillna(0)
            dfB048['TonKD']=dfB048['TongTon']-dfB048['SLYC']
            df=pd.merge(df,dfB048,on='MaHang',how='left')
            df=df.fillna(0)
            df=df[df['TonKD']<df['SS']]
            df['Ty le']=df['TonKD']/df['SS']
            df["Ty le"] = df["Ty le"].apply(lambda x: x * 100)
            df["Ty le"] = df["Ty le"].round(1)
            df["SS"] = df["SS"].round(1)
            df=df.sort_values(by='Ty le')
            df['TonKD']=np.where(df['TonKD']<0,0,df['TonKD'])
            df['Ty le']=np.where(df['TonKD']<0,0,df['Ty le'])
            dfSS=df
            df1=df[['MaHang','TenHang','TonKD','SS','Ty le']]
            ###############
            df1=df1[df1['TonKD']>0]
            #####################
            self.ui.safetyStock.setRowCount(len(df1))
            for row_number, (_,row_data) in enumerate(df1.iterrows()):
                for column_number, data in enumerate(row_data):
                    item = QtWidgets.QTableWidgetItem(str(data))
                    self.ui.safetyStock.setItem(row_number, column_number, item)
            self.ui.safetyStock.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
            self.ui.safetyStock.setColumnWidth(0, 150)
            self.ui.safetyStock.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
            self.ui.safetyStock.setColumnWidth(1, 490)
            self.ui.safetyStock.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
            self.ui.safetyStock.setColumnWidth(2, 230)
            self.ui.safetyStock.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
            self.ui.safetyStock.setColumnWidth(3, 220)
            self.ui.safetyStock.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.Fixed)
            self.ui.safetyStock.setColumnWidth(4, 220)
            delegate = AlignRDelegate()
            self.ui.safetyStock.setItemDelegateForColumn(2, delegate)
            self.ui.safetyStock.setItemDelegateForColumn(3, delegate)
            self.ui.safetyStock.setItemDelegateForColumn(4, delegate)
            return dfSS
    def SSAi(self):
        # df=pd.read_excel(r'D:\IMIS\Dashboard\ss.xlsx')
        global dfSS
        global dfABC
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        c.execute("select LichSuGD.MaHang,TenHang,NgayXK, sum(SLXuat) as NhuCau from LichSuGD,HangHoa where LichSuGD.MaHang=HangHoa.MaHang group by LichSuGD.MaHang,TenHang,DonGia,NgayXK")
        df1=c.fetchall()
        data =[]
        for i in df1:
            i=tuple(i)
            data.append(i)
        names = [ x[0] for x in c.description]
        df1 = pd.DataFrame(data, columns=names)
        with open('models.pkl', 'rb') as f:
            models = pickle.load(f)
        def split_sequence(sequence, n_steps):
            X, y = list(), list()
            for i in range(len(sequence)):
                # find the end of this pattern
                end_ix = i + n_steps
                # check if we are beyond the sequence
                if end_ix > len(sequence)-1:
                    break
                # gather input and output parts of the pattern
                seq_x, seq_y = sequence[i:end_ix], sequence[end_ix]
                X.append(seq_x)
                y.append(seq_y)
            return array(X), array(y)
        mh=[]
        tbinh=[]
        dlc=[]
        min=pd.to_datetime(date(2022,1,1))
        max=pd.to_datetime(date(2022,12,31))
        gtrilon=list(models.keys())
        # Tạo một chuỗi ngày tháng liên tục từ ngày nhỏ nhất đến ngày lớn nhất
        date_range = pd.date_range(start=min, end=max)
        # Tạo một DataFrame mới với cột ngày tháng liên tục
        new_df = pd.DataFrame({'NgayXK': date_range})
        for name, group in df1.groupby(["MaHang"]):
            #Sử dụng hàm merge để kết hợp hai DataFrame
            mh.append(name[0])
            group['NgayXK']=pd.to_datetime(group['NgayXK'])
            group = pd.merge(new_df, group, on='NgayXK', how='left')
            group['NhuCau'] = group['NhuCau'].fillna(0)
            group['Month']=group['NgayXK'].dt.month
            group = group.groupby(['Month']).agg({'NhuCau': 'sum'}).reset_index()
            # define input sequence
            raw_seq = group['NhuCau'].values
            if name in gtrilon:
                model=models.get(name)
                n_steps = 3
                # split into samples
                X, y = split_sequence(raw_seq, n_steps)
                #demonstrate prediction for the 12 months
                x_input=X[-1].reshape((3,))
                temp_input=list(x_input)
                list_output=[]
                i=0
                while (i<12):
                    if len(temp_input)>3:
                        x_input=array(temp_input[1:])
                        x_input=x_input.reshape((1,n_steps,1))
                        yhat=model.predict(x_input,verbose=0)
                        temp_input.append(yhat[0][0])
                        temp_input=temp_input[1:]
                        list_output.append(yhat[0][0])
                        i+=1
                    else:
                        x_input=x_input.reshape((1,n_steps,1))
                        yhat=model.predict(x_input,verbose=0)
                        temp_input.append(yhat[0][0])
                        list_output.append(yhat[0][0])
                        i+=1
                    # print(list_output)
                tb=np.round(np.average(list_output)/24,2)
                dl=np.round(np.std(list_output)/24,2)
                tbinh.append(tb)
                dlc.append(dl)
            else:
                tb=np.round(np.average(raw_seq)/24,2)
                dl=np.round(np.std(raw_seq)/24,2)
                tbinh.append(tb)
                dlc.append(dl)
        # a = df1.groupby(['MaHang']).NhuCau.agg(['mean', lambda x: x.std(ddof=0)])
        # a.columns=['mean','std']
        # a['std']=a['std'].apply(lambda x: round(x, 2))
        # a['mean']=a['mean'].apply(lambda x: round(x, 2))
        # a=a.reset_index()
        datadf = {
            'MaHang': mh,
            'mean': tbinh,
            'std': dlc}
        a = pd.DataFrame(datadf)
        print(a)
        a['std']=np.where(a['std']==0,1,a['std'])
        dfABC_SS= dfABC
        df=pd.merge(a,dfABC_SS,on='MaHang',how='inner')
        print(df)
        # setup service level
        a=self.ui.sla.toPlainText()
        b=self.ui.slb.toPlainText()
        d=self.ui.slc.toPlainText()
        if a == '':
            return 0.0  # or any other default value you want to assign
        else:
            a=float(a)/100
        if b == '':
            return 0.0  # or any other default value you want to assign
        else:
            b=float(b)/100
        if d == '':
            return 0.0  # or any other default value you want to assign
        else:
            d=float(d)/100
        df['ServiceLevel']=np.where(df['Class']=='A',a,np.where(df['Class']=='B',b,d))
        df['R']= np.round(norm.ppf(df['ServiceLevel'],df['mean'],df['std']),decimals=2)
        df['z']=(df['R']-df['mean'])/df['std']
        df['SS']=df['z']*df['std']
        c.execute("select MaHang,TongTon from B048")
        dfB048=c.fetchall()
        data1 =[]
        for i in dfB048:
            i=tuple(i)
            data1.append(i)
        names1 = [ x[0] for x in c.description]
        dfB048 = pd.DataFrame(data1, columns=names1)
        c.execute("select MaHang,SLYC from A040Detail")
        dfYCGH=c.fetchall()
        data2 =[]
        for i in dfYCGH:
            i=tuple(i)
            data2.append(i)
        names2 = [ x[0] for x in c.description]
        dfYCGH = pd.DataFrame(data2, columns=names2)
        dfYCGH=dfYCGH.groupby(['MaHang'])['SLYC'].sum()
        dfYCGH=dfYCGH.reset_index()
        # print(dfYCGH)
        dfB048=dfB048.groupby(['MaHang'])['TongTon'].sum()
        dfB048=dfB048.reset_index()
        # print(dfB048)
        dfB048=dfB048.merge(dfYCGH,on='MaHang',how='left')
        dfB048=dfB048.fillna(0)
        dfB048['TonKD']=dfB048['TongTon']-dfB048['SLYC']
        df=pd.merge(df,dfB048,on='MaHang',how='left')
        df=df.fillna(0)
        df=df[df['TonKD']<df['SS']]
        df['Ty le']=df['TonKD']/df['SS']
        df["Ty le"] = df["Ty le"].apply(lambda x: x * 100)
        df["Ty le"] = df["Ty le"].round(1)
        df["SS"] = df["SS"].round(1)
        df=df.sort_values(by='Ty le')
        dfSS=df
        df1=df[['MaHang','TenHang','TonKD','SS','Ty le']]
        ###############
        df1=df1[df1['TonKD']>0]
        #####################
        self.ui.safetyStock.setRowCount(len(df1))
        for row_number, (_,row_data) in enumerate(df1.iterrows()):
            for column_number, data in enumerate(row_data):
                item = QtWidgets.QTableWidgetItem(str(data))
                self.ui.safetyStock.setItem(row_number, column_number, item)
        self.ui.safetyStock.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
        self.ui.safetyStock.setColumnWidth(0, 150)
        self.ui.safetyStock.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
        self.ui.safetyStock.setColumnWidth(1, 490)
        self.ui.safetyStock.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
        self.ui.safetyStock.setColumnWidth(2, 230)
        self.ui.safetyStock.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
        self.ui.safetyStock.setColumnWidth(3, 220)
        self.ui.safetyStock.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.Fixed)
        self.ui.safetyStock.setColumnWidth(4, 220)
        delegate = AlignRDelegate()
        self.ui.safetyStock.setItemDelegateForColumn(2, delegate)
        self.ui.safetyStock.setItemDelegateForColumn(3, delegate)
        self.ui.safetyStock.setItemDelegateForColumn(4, delegate)
        return dfSS
    def ROP(self):
        # global dfSS
        def rop_button_clicked():
            # df = dfSS
            df=AppFunction.SS(self)
            conn=AppFunction.create_connection(self)
            c=conn.cursor()
            c.execute("select MaHang,LT from HangHoa")
            dfHH=c.fetchall()
            data =[]
            for i in dfHH:
                i=tuple(i)
                data.append(i)
            names = [ x[0] for x in c.description]
            dfHH = pd.DataFrame(data, columns=names)
            df=pd.merge(df,dfHH,on='MaHang',how='inner')
            df['LTD']=df['LT']*df['mean']
            df['ROP']=df['SS']+df['LTD']
            # print(df)
            # so luong SR chua lay tu ncc
            c.execute("select MaHang,SLChuaGiao,NgayDuyetDH from A010,A010Detail where A010.MaDH=A010Detail.MaDH")
            dfDH=c.fetchall()
            data1 =[]
            for i in dfDH:
                i=tuple(i)
                data1.append(i)
            names1 = [ x[0] for x in c.description]
            dfDH = pd.DataFrame(data1, columns=names1)
            dfDH=dfDH.groupby(['MaHang','NgayDuyetDH'])['SLChuaGiao'].sum()
            dfDH=dfDH.reset_index() # SLChuaGiao is SR
            # so luong BO cn dat chua lay chinh la SLYC
            # tinh ip = TonKD+SR
            df=pd.merge(df,dfDH,on='MaHang',how='left')
            df=df.rename(columns={'SLChuaGiao':'SR'})
            df['IP']=df['TonKD']+df['SR']
            df=df.fillna(0)
            # convert the date columns to datetime format
            tz_VN = pytz.timezone('Asia/Ho_Chi_Minh')
            datetime_VN = datetime.now(tz_VN)
            now_vn = datetime_VN.strftime("%d/%m/%Y")
            df['end_date']=now_vn
            df['start_date'] = pd.to_datetime(df['NgayDuyetDH'])
            df['end_date'] = pd.to_datetime(df['end_date'])
            # calculate the number of days between the two columns
            df['days_between'] = (df['end_date'] - df['start_date']).dt.days
            # so sánh TongTon vs ROP=SS+LTD
            df['Danh gia']=np.where(df['TonKD']<df['ROP'],'Mua','Không mua')
            # xét chi tiết IP trong trường hợp mua
            df['Danh gia']=np.where((df['TonKD']<df['ROP'])&(df['IP']<df['ROP']),'Mua','Không mua')
            # có phải khi IP>LTD là luôn không phải mua? NO
            # IP = TongTon+SR-BO vs ROP
            # df['TL co the sd']=(df['IP']-df['ROP'])/df['mean']
            # df['days']=df['LT']-df['days_between']
            # df['So sanh']=df['days']-math.floor(df['TL co the sd'])
            df['Danh gia']=np.where((df['TonKD']<df['ROP'])&(df['IP']>=df['ROP']) & (df['days_between']!=df['LT']),'Mua',df['Danh gia']) # chỉ xét tại thời điểm báo cáo
            # xet
            df=df[df['Danh gia']=='Mua']
            ###############
            df=df[df['TonKD']>0]
            #####################
            df=df[["MaHang","TenHang","TonKD",'mean']]
            self.ui.dsHHCanMua.setRowCount(len(df))
            for row_number, (_,row_data) in enumerate(df.iterrows()):
                for column_number, data in enumerate(row_data):
                    item = QtWidgets.QTableWidgetItem(str(data))
                    self.ui.dsHHCanMua.setItem(row_number, column_number, item)
            self.ui.dsHHCanMua.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
            self.ui.dsHHCanMua.setColumnWidth(0, 150)
            self.ui.dsHHCanMua.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
            self.ui.dsHHCanMua.setColumnWidth(1, 700)
            self.ui.dsHHCanMua.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
            self.ui.dsHHCanMua.setColumnWidth(2, 230)
            self.ui.dsHHCanMua.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
            self.ui.dsHHCanMua.setColumnWidth(3, 220)
            delegate = AlignRDelegate()
            self.ui.dsHHCanMua.setItemDelegateForColumn(2, delegate)
            self.ui.dsHHCanMua.setItemDelegateForColumn(3, delegate)
            self.ui.dsHHCanMua.itemSelectionChanged.connect(lambda: AppFunction.on_table_item_selection_changed(self))
        self.ui.rop.clicked.connect(rop_button_clicked)
    def on_table_item_selection_changed(self):
        self.ui.mainPages.setCurrentIndex(14)  # Chuyển đến trang 29
        selected_indexes = self.ui.dsHHCanMua.selectedIndexes()
        if selected_indexes:
            row = selected_indexes[0].row()
            values = []
            for column in range(self.ui.dsHHCanMua.columnCount()):
                item = self.ui.dsHHCanMua.item(row, column)
                if item is not None:
                    values.append(item.text())
            # print(values)
        mh=values[0]
        th=values[1]
        demand=values[3]
        self.ui.mh.setText(mh)
        self.ui.th.setText(th)
        self.ui.nhuCau.setText(str(demand))
        AppFunction.eoq_mh.append(mh)
        AppFunction.eoq_th.append(th)
        # print(AppFunction.eoq_mh)
    def EOQ(self):
        cplk=self.ui.chiPhiLk.toPlainText() # H chi phí lưu kho/tháng
        cpdh=self.ui.chiPhiDh.toPlainText() #S chi phí một lần đặt hàng
        ncau=self.ui.nhuCau.toPlainText() #D
        cplk=float(cplk)
        cpdh=float(cpdh)
        ncau=float(ncau)
        # nhu cầu cả năm, giả định công ty làm việc 52 tuần, mỗi tuần làm việc 6 ngày
        D=ncau*52*6
        # chi phí lưu kho 1 năm
        H=cplk*12
        # chi phí 1 lần đặt hàng
        S=cpdh
        # tính eoq
        eoq=np.sqrt((2*D*S)/H)
        eoq=np.round(eoq,1)
        self.ui.kq.setText(str(eoq))
        AppFunction.eoq_q.append(eoq)
    def Back(self):
        self.ui.mainPages.setCurrentIndex(13)
        self.ui.chiPhiLk.setText("")
        self.ui.chiPhiDh.setText("")
        self.ui.kq.setText("")
    def BuildModel(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        c.execute("select * from LichSuGD")
        dfLSGD=c.fetchall()
        data =[]
        for i in dfLSGD:
            i=tuple(i)
            data.append(i)
        names = [ x[0] for x in c.description]
        dfLSGD = pd.DataFrame(data, columns=names)
        c.execute("select MaHang,DonGia from HangHoa")
        dfHH=c.fetchall()
        data1 =[]
        for i in dfHH:
            i=tuple(i)
            data1.append(i)
        names1 = [ x[0] for x in c.description]
        dfHH = pd.DataFrame(data1, columns=names1)
        dfLSGD=dfLSGD.merge(dfHH,on='MaHang',how='inner')
        df1 = dfLSGD[dfLSGD['DonGia'].astype(str).str.len() == 9]
        current_year = datetime.now().year
        df1['NgayXK']=pd.to_datetime(df1['NgayXK'])
        df1 = df1[df1['NgayXK'].dt.year < current_year]
        # df1=df1.rename(columns={'Số lượng':'Demand','Ngày tính tồn kho':'Date'})
        # Chuyển đổi cột ngày tháng sang định dạng datetime
        # df1['Date'] = pd.to_datetime(df1['Date'])
        # split a univariate sequence
        def split_sequence(sequence, n_steps):
            X, y = list(), list()
            for i in range(len(sequence)):
                # find the end of this pattern
                end_ix = i + n_steps
                # check if we are beyond the sequence
                if end_ix > len(sequence)-1:
                    break
                # gather input and output parts of the pattern
                seq_x, seq_y = sequence[i:end_ix], sequence[end_ix]
                X.append(seq_x)
                y.append(seq_y)
            return array(X), array(y)
        # Define a list to store your models
        models1 = {}
        min=pd.to_datetime(date(2022,1,1))
        max=pd.to_datetime(date(2022,12,31))
        # Tạo một chuỗi ngày tháng liên tục từ ngày nhỏ nhất đến ngày lớn nhất
        date_range = pd.date_range(start=min, end=max)
        # Tạo một DataFrame mới với cột ngày tháng liên tục
        new_df = pd.DataFrame({'NgayXK': date_range})
        for name, group in df1.groupby(["MaHang"]):
            # Sử dụng hàm merge để kết hợp hai DataFrame
            group = pd.merge(new_df, group, on='NgayXK', how='left')
            group['SLXuat'] = group['SLXuat'].fillna(0)
            group['Month']=group['NgayXK'].dt.month
            group = group.groupby(['Month']).agg({'SLXuat': 'sum'}).reset_index()
            detrended = detrend(group['SLXuat'], type='linear')
            detrended = pd.Series(detrended, index=group.index)
            # define input sequence
            raw_seq = detrended
            # choose a number of time steps
            n_steps = 3
            # split into samples
            X, y = split_sequence(raw_seq, n_steps)
            q_80 = int(len(X) * .8)
            X_train, y_train = X[:q_80], y[:q_80]
            X_test, y_test = X[q_80:], y[q_80:]
            scaler = StandardScaler()
            X_train = scaler.fit_transform(X_train)
            X_test = scaler.fit_transform(X_test)
            # reshape from [samples, timesteps] into [samples, timesteps, features]
            n_features = 1
            X_train = X_train.reshape((X_train.shape[0], X_train.shape[1], n_features))
            X_test = X_test.reshape((X_test.shape[0], X_test.shape[1], n_features))
            # define model
            model = Sequential()
            model.add(Bidirectional(LSTM(50, activation='relu'), input_shape=(n_steps, n_features)))
            model.add(Dense(1))
            model.compile(optimizer='adam', loss='mse')
            # fit model
            model.fit(X_train, y_train, epochs=200, verbose=0)
            # Append the trained model to the list
            models1[name]=model
        # Save the list of models using pickle
        with open('models.pkl', 'wb') as f:
            pickle.dump(models1, f)
        QMessageBox.information(self, "Thông báo", f"Build model thành công!")
    def DatHangBC(self):
        mh=set(AppFunction.eoq_mh)
        th=set(AppFunction.eoq_th)
        eoq=set(AppFunction.eoq_q)
        df = pd.DataFrame({'mh': list(mh), 'th': list(th), 'eoq': list(eoq)})
        df=df.rename(columns={'mh':'MaHang'})
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        c.execute("select MaHang,DonGia,TenNhomNH from HangHoa,NhomNH where HangHoa.MaNhomNH=NhomNH.MaNhomNH")
        dfHH=c.fetchall()
        data1 =[]
        for i in dfHH:
            i=tuple(i)
            data1.append(i)
        names1 = [ x[0] for x in c.description]
        dfHH = pd.DataFrame(data1, columns=names1)
        df=df.merge(dfHH,on='MaHang',how='inner')
        df['GiaTri']=np.round(df['DonGia']*df['eoq'],1)
        df=df[['TenNhomNH','MaHang','th','eoq','GiaTri']]
        data=df.values.tolist()
        col=['Nhóm ngành hàng','Mã hàng','Tên hàng','Số lượng đặt','Giá trị']
        pdf = DatHang()
        pdf.add_page()
        pdf.alias_nb_pages()
        pdf.set_auto_page_break(False)
        pdf.set_widths([40, 30, 60,20,40])
        pdf.set_font('DejaVu',size=10)
        pdf.set_x(10)
        pdf.multi_cell(0, 5,'Công ty Cổ phần Viễn thông FPT \nTrung tâm Phát triển và Quản lý hạ tầng MB \nĐịa chỉ: 48 Vạn Bảo, Ngọc Khánh, Ba Đình', border="B", align='L')
        pdf.ln(5) 
        pdf.set_font('DejaVu', 'B', 15)
        pdf.multi_cell(0, 7, 'BÁO CÁO SỐ LƯỢNG ĐẶT HÀNG', border=0, align='C')
        pdf.ln(3)
        pdf.set_font('DejaVu', '', 10)
        tz_VN = pytz.timezone('Asia/Ho_Chi_Minh')
        datetime_VN = datetime.now(tz_VN)
        now_vn = datetime_VN.strftime("%d/%m/%Y")
        pdf.multi_cell(0, 7, f'Ngày: {now_vn}', border=0, align='R')
        pdf.ln(5)
        pdf.tieude(col) 
        for i in range(len(data)):
            if pdf.get_y() + 30 > pdf.h:
                pdf.add_page()
                pdf.tieude(col)
            pdf.row([data[i]])
        sum1=df["GiaTri"].sum()
        tong=["Tổng cộng",sum1]
        pdf.sum(tong)
        # pdf.row(data)
        filename = filedialog.asksaveasfilename(defaultextension='.pdf')
        if filename:
            pdf.output(filename, 'F')
        # file_path = filedialog.asksaveasfilename(defaultextension='.xlsx')
        # if file_path:
        #     writer = pd.ExcelWriter(file_path,engine="openpyxl")
        #     df.to_excel(writer, index=False)
    def BCABC(self):
        # cb=self.ui.canhBaoCb.currentText()
        # if cb!="Tất cả":
        #     df=AppFunction.FilterCanhBao(self)
        # else:
        #     df=AppFunction.Alert(self)
        global dfABC
        df=dfABC
        df=df[["MaHang","TenHang","DonGia","Giá trị","RunPerc","Class"]]
        # df=pd.read_excel(r'D:\IMIS\Dashboard\abc.xlsx')
        # df["Percent"] = df["Percent"].apply(lambda x: x * 100)
        # df["Percent"] = df["Percent"].round(1)
        data=df.values.tolist()
        col=['Mã hàng','Tên hàng','Đơn giá','Giá trị hàng năm','% về giá trị so với tổng giá trị năm','Xếp loại']
        pdf = ABC()
        pdf.add_page()
        pdf.alias_nb_pages()
        pdf.set_auto_page_break(False)
        pdf.set_widths([25, 40, 30,30,30,35])
        pdf.set_font('DejaVu',size=10)
        pdf.set_x(10)
        pdf.multi_cell(0, 5,'Công ty Cổ phần Viễn thông FPT \nTrung tâm Phát triển và Quản lý hạ tầng MB \nĐịa chỉ: 48 Vạn Bảo, Ngọc Khánh, Ba Đình', border="B", align='L')
        pdf.ln(5) 
        pdf.set_font('DejaVu', 'B', 15)
        pdf.multi_cell(0, 7, 'BÁO CÁO KẾT QUẢ PHÂN TÍCH ABC', border=0, align='C')
        pdf.ln(3)
        pdf.set_font('DejaVu', '', 10)
        tz_VN = pytz.timezone('Asia/Ho_Chi_Minh')
        datetime_VN = datetime.now(tz_VN)
        now_vn = datetime_VN.strftime("%d/%m/%Y")
        pdf.multi_cell(0, 7, f'Ngày: {now_vn}', border=0, align='R')
        pdf.ln(5)
        pdf.tieude(col) 
        for i in range(len(data)):
            if pdf.get_y() + 40 > pdf.h:
                pdf.add_page()
                pdf.tieude(col)
            # pdf.add_page()
            # pdf.tieude(col)
            pdf.row([data[i]])
            # pdf.row(data)
        filename = filedialog.asksaveasfilename(defaultextension='.pdf')
        if filename:
            pdf.output(filename, 'F')
    def BCSS(self):
        # cb=self.ui.canhBaoCb.currentText()
        # if cb!="Tất cả":
        #     df=AppFunction.FilterCanhBao(self)
        # else:
        #     df=AppFunction.Alert(self)
        global dfSS
        df=dfSS
        df=df[['MaHang','TenHang','TonKD','SS','Ty le']]
        # df=pd.read_excel(r'D:\IMIS\Dashboard\ss.xlsx')
        # df["Chênh lệch"] = df["Chênh lệch"].apply(lambda x: x * 100)
        # df["Chênh lệch"] = df["Chênh lệch"].round(1)
        data=df.values.tolist()
        col=['Mã hàng','Tên hàng','Số lượng tồn','Mức tồn kho an toàn','% Chênh lệch']
        pdf = SS()
        pdf.set_auto_page_break(False)
        pdf.add_page()
        pdf.alias_nb_pages()
        # pdf.set_auto_page_break(False)
        pdf.set_widths([25, 55, 40,40,30])
        pdf.set_font('DejaVu',size=10)
        pdf.set_x(10)
        pdf.multi_cell(0, 5,'Công ty Cổ phần Viễn thông FPT \nTrung tâm Phát triển và Quản lý hạ tầng MB \nĐịa chỉ: 48 Vạn Bảo, Ngọc Khánh, Ba Đình', border="B", align='L')
        pdf.ln(5) 
        pdf.set_font('DejaVu', 'B', 15)
        pdf.multi_cell(0, 7, 'BÁO CÁO CẢNH BÁO TỒN KHO', border=0, align='C')
        pdf.ln(3)
        pdf.set_font('DejaVu', '', 10)
        tz_VN = pytz.timezone('Asia/Ho_Chi_Minh')
        datetime_VN = datetime.now(tz_VN)
        now_vn = datetime_VN.strftime("%d/%m/%Y")
        pdf.multi_cell(0, 7, f'Ngày: {now_vn}', border=0, align='R')
        pdf.ln(5)
        pdf.tieude(col) 
        for i in range(len(data)):
            if pdf.get_y() + 30 > pdf.h:
                pdf.add_page()
                pdf.tieude(col)
            pdf.row([data[i]])
        # pdf.row(data)
        filename = filedialog.asksaveasfilename(defaultextension='.pdf')
        if filename:
            pdf.output(filename, 'F')
    def BCTK(self):
        global dfTK
        df=dfTK
        data=df.values.tolist()
        col=['Tên tài khoản','Mật khẩu','Loại tài khoản']
        pdf = TK()
        pdf.set_auto_page_break(False)
        pdf.add_page()
        pdf.alias_nb_pages()
        # pdf.set_auto_page_break(False)
        pdf.set_widths([50,50,70])
        pdf.set_font('DejaVu',size=10)
        pdf.set_x(10)
        pdf.multi_cell(0, 5,'Công ty Cổ phần Viễn thông FPT \nTrung tâm Phát triển và Quản lý hạ tầng MB \nĐịa chỉ: 48 Vạn Bảo, Ngọc Khánh, Ba Đình', border="B", align='L')
        pdf.ln(5) 
        pdf.set_font('DejaVu', 'B', 15)
        pdf.multi_cell(0, 7, 'BÁO CÁO DANH SÁCH TÀI KHOẢN', border=0, align='C')
        pdf.ln(3)
        pdf.set_font('DejaVu', '', 10)
        tz_VN = pytz.timezone('Asia/Ho_Chi_Minh')
        datetime_VN = datetime.now(tz_VN)
        now_vn = datetime_VN.strftime("%d/%m/%Y")
        pdf.multi_cell(0, 7, f'Ngày: {now_vn}', border=0, align='R')
        pdf.ln(5)
        pdf.tieude(col) 
        for i in range(len(data)):
            if pdf.get_y() + 30 > pdf.h:
                pdf.add_page()
                pdf.tieude(col)
            pdf.row([data[i]])
        # pdf.row(data)
        filename = filedialog.asksaveasfilename(defaultextension='.pdf')
        if filename:
            pdf.output(filename, 'F')
    def Alert(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        c.execute("select MaHang,sum(SLTonKD) as SLTonKD from B048 group by MaHang order by sum(SLTonKD) desc")
        dfB048=c.fetchall()
        data =[]
        for i in dfB048:
            i=tuple(i)
            data.append(i)
        names = [ x[0] for x in c.description]
        dfB048 = pd.DataFrame(data, columns=names)
        c.execute("select MaHang,sum(SLChuaGiao) as SLChuaGiao from A010 group by MaHang order by SLChuaGiao desc")
        dfDH=c.fetchall()
        data2 =[]
        for i in dfDH:
            i=tuple(i)
            data2.append(i)
        names2 = [ x[0] for x in c.description]
        dfDH = pd.DataFrame(data2, columns=names2)
        c.execute("select MaHang,sum(SLYC) as SLYC from A040 group by MaHang order by sum(SLYC) desc")
        dfYCGH=c.fetchall()
        data3 =[]
        for i in dfYCGH:
            i=tuple(i)
            data3.append(i)
        names3 = [ x[0] for x in c.description]
        dfYCGH = pd.DataFrame(data3, columns=names3)
        c.execute("select MaHang,TenHang,SLMin from HangHoa where SLMin is not null")
        dfHH=c.fetchall()
        data4 =[]
        for i in dfHH:
            i=tuple(i)
            data4.append(i)
        names4 = [ x[0] for x in c.description]
        dfHH = pd.DataFrame(data4, columns=names4)
        df=pd.merge(dfB048,dfDH,on='MaHang',how='inner')
        df=pd.merge(df,dfYCGH,on='MaHang',how='inner')
        df=pd.merge(df,dfHH,on='MaHang',how='right')
        df.fillna(0)
        df["Cảnh báo"]=df.apply(AppFunction.warning_level,axis=1)
        df=df[["MaHang","TenHang","SLTonKD","SLYC","SLChuaGiao","Cảnh báo"]]
        # df = df.style.applymap(AppFunction.color_cell, subset=["Cảnh báo"])
        # df = df.data
        self.ui.canhBaoTb.setRowCount(len(df))
        for row_number, (_,row_data) in enumerate(df.iterrows()):
            for column_number, data in enumerate(row_data):
                item = QtWidgets.QTableWidgetItem(str(data))
                # if df.loc[row_number, "Cảnh báo"] == "Cấp 1":
                #     item.setBackground(QColor("red"))
                # elif df.loc[row_number, "Cảnh báo"] == "Cấp 2":
                #     item.setBackground(QColor("orange"))
                # elif df.loc[row_number, "Cảnh báo"] == "Cấp 3":
                #     item.setBackground(QColor("yellow"))
                # else:
                #     item.setBackground(QColor("green"))
                if data == "Cấp 1":
                    item.setBackground(QColor("red"))
                elif data == "Cấp 2":
                    item.setBackground(QColor("orange"))
                elif data == "Cấp 3":
                    item.setBackground(QColor("yellow"))
                elif data == "Bình thường":
                    item.setBackground(QColor("green"))
                self.ui.canhBaoTb.setItem(row_number, column_number, item)
        self.ui.canhBaoTb.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
        self.ui.canhBaoTb.setColumnWidth(0, 120)
        # Set the width of the third column to fill the remaining space
        self.ui.canhBaoTb.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
        self.ui.canhBaoTb.setColumnWidth(1, 400)
        self.ui.canhBaoTb.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
        self.ui.canhBaoTb.setColumnWidth(2, 150)
        self.ui.canhBaoTb.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
        self.ui.canhBaoTb.setColumnWidth(3, 150)
        self.ui.canhBaoTb.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.Fixed)
        self.ui.canhBaoTb.setColumnWidth(4, 150)
        self.ui.canhBaoTb.horizontalHeader().setSectionResizeMode(5, QtWidgets.QHeaderView.Fixed)
        self.ui.canhBaoTb.setColumnWidth(5, 150)
        delegate = AlignRDelegate()
        self.ui.canhBaoTb.setItemDelegateForColumn(2, delegate)
        self.ui.canhBaoTb.setItemDelegateForColumn(3, delegate)
        self.ui.canhBaoTb.setItemDelegateForColumn(4, delegate)
        self.ui.canhBaoTb.setItemDelegateForColumn(5, delegate)
        return df
    def HoanTacCanhBao(self):
        AppFunction.Alert(self)
        self.ui.canhBaoCb.setCurrentIndex(0)
    def FilterCanhBao(self):
        conn=AppFunction.create_connection(self)
        c=conn.cursor()
        cb=self.ui.canhBaoCb.currentText()
        df=AppFunction.Alert(self)
        # df=df.query('"Cảnh báo"==cb')
        df=df[df["Cảnh báo"]==cb]
        self.ui.canhBaoTb.setRowCount(len(df))
        for row_number, (_,row_data) in enumerate(df.iterrows()):
            for column_number, data in enumerate(row_data):
                item = QtWidgets.QTableWidgetItem(str(data))
                if data == "Cấp 1":
                    item.setBackground(QColor("red"))
                elif data == "Cấp 2":
                    item.setBackground(QColor("orange"))
                elif data == "Cấp 3":
                    item.setBackground(QColor("yellow"))
                elif data == "Bình thường":
                    item.setBackground(QColor("green"))
                self.ui.canhBaoTb.setItem(row_number, column_number, item)
        self.ui.canhBaoTb.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
        self.ui.canhBaoTb.setColumnWidth(0, 120)
        # Set the width of the third column to fill the remaining space
        self.ui.canhBaoTb.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
        self.ui.canhBaoTb.setColumnWidth(1, 400)
        self.ui.canhBaoTb.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
        self.ui.canhBaoTb.setColumnWidth(2, 150)
        self.ui.canhBaoTb.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
        self.ui.canhBaoTb.setColumnWidth(3, 150)
        self.ui.canhBaoTb.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.Fixed)
        self.ui.canhBaoTb.setColumnWidth(4, 150)
        self.ui.canhBaoTb.horizontalHeader().setSectionResizeMode(5, QtWidgets.QHeaderView.Fixed)
        self.ui.canhBaoTb.setColumnWidth(5, 150)
        delegate = AlignRDelegate()
        self.ui.canhBaoTb.setItemDelegateForColumn(2, delegate)
        self.ui.canhBaoTb.setItemDelegateForColumn(3, delegate)
        self.ui.canhBaoTb.setItemDelegateForColumn(4, delegate)
        self.ui.canhBaoTb.setItemDelegateForColumn(5, delegate)
        return df
    def BCCanhBao(self):
        cb=self.ui.canhBaoCb.currentText()
        if cb!="Tất cả":
            df=AppFunction.FilterCanhBao(self)
        else:
            df=AppFunction.Alert(self)
        data=df.values.tolist()
        col=['Mã hàng','Tên hàng','SL tồn khả dụng','SL tồn YCGH','SL tồn đơn hàng','Cảnh báo']
        pdf = CanhBao()
        pdf.add_page()
        pdf.alias_nb_pages()
        pdf.set_auto_page_break(False)
        pdf.set_widths([25, 40, 25,25,25,50])
        pdf.set_font('DejaVu',size=10)
        pdf.set_x(10)
        pdf.multi_cell(0, 5,'Công ty Cổ phần Viễn thông FPT \nTrung tâm Phát triển và Quản lý hạ tầng MB \nĐịa chỉ: 48 Vạn Bảo, Ngọc Khánh, Ba Đình', border="B", align='L')
        pdf.ln(5) 
        pdf.set_font('DejaVu', 'B', 15)
        pdf.multi_cell(0, 7, 'BÁO CÁO CẢNH BÁO TỒN KHO', border=0, align='C')
        pdf.ln(3)
        pdf.set_font('DejaVu', '', 10)
        tz_VN = pytz.timezone('Asia/Ho_Chi_Minh')
        datetime_VN = datetime.now(tz_VN)
        now_vn = datetime_VN.strftime("%d/%m/%Y")
        pdf.multi_cell(0, 7, f'Ngày: {now_vn}', border=0, align='R')
        pdf.ln(5)
        pdf.tieude(col) 
        for i in range(len(data)):
            if pdf.get_y() + 30 > pdf.h:
                pdf.add_page()
                pdf.tieude(col)
            pdf.row([data[i]])
        # pdf.row(data)
        filename = filedialog.asksaveasfilename(defaultextension='.pdf')
        if filename:
            pdf.output(filename, 'F')
class PDFB048(FPDF):
    def __init__(self):
        super().__init__()
        # self.widths = [40, 60, 80]
        self.add_font('DejaVu','', r'ttf\DejaVuSansCondensed.ttf', uni=True)
        self.add_font('DejaVu','B', r'ttf\DejaVuSansCondensed-Bold.ttf', uni=True)
    def set_widths(self, widths):
        self.widths = widths
    def sum(self,data):
        self.widths=[145,22.5,22.5]
        for i, item in enumerate(data):
            x = self.get_x()
            y = self.get_y()
            self.set_font('DejaVu',style="B",size=11)
            self.multi_cell(self.widths[i], 7, str(item), border=1,align="C")
            self.set_xy(x+self.widths[i],y)
    def tieude(self,data):
        max_height = 0
        a=[10, 30, 55,115,155,177.5,200]
        for i, item in enumerate(data):
            x = self.get_x()
            y = self.get_y()
            self.set_font('DejaVu',style="B",size=11)
            self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
            if self.y - y > max_height:
                max_height = self.y - y
            self.set_xy(x+self.widths[i],y)
            self.line(x, y, x+self.widths[i], y)
            self.line(x, y, x, y+max_height)
        for x in a:
            self.line(x,y,x,y+max_height)
        self.ln(max_height)
        x=10
        self.line(x, self.get_y(), x+190, self.get_y())
    def footer(self):
        # Set position of the footer
        self.set_y(-10)
        # Set font size
        self.set_font('DejaVu', '', 8)
        # Add page number
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', 0, 0, 'C')
    def row(self, data):
        for row in data:
            max_height = 0
            a=[10, 30, 55,115,155,177.5,200]
            for i, item in enumerate(row):
                x = self.get_x()
                y = self.get_y()
                self.set_font('DejaVu',style="",size=11)
                self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
                if self.y - y > max_height:
                    max_height = self.y - y
                self.set_xy(x+self.widths[i],y)
                self.line(x, y, x+self.widths[i], y) #horizontal line
            self.ln(max_height)
            for x in a:
                self.line(x,y,x,y+max_height)
        x=10
        self.line(x, self.get_y(), x+190, self.get_y())
class CanhBao(FPDF):
    def __init__(self):
        super().__init__()
        # self.widths = [40, 60, 80]
        self.add_font('DejaVu','', r'ttf\DejaVuSansCondensed.ttf', uni=True)
        self.add_font('DejaVu','B', r'ttf\DejaVuSansCondensed-Bold.ttf', uni=True)
    def set_widths(self, widths):
        self.widths = widths
    def tieude(self,data):
        max_height = 0
        a=[10, 35, 75,100,125,150,200]
        for i, item in enumerate(data):
            x = self.get_x()
            y = self.get_y()
            self.set_font('DejaVu',style="B",size=11)
            self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
            if self.y - y > max_height:
                max_height = self.y - y
            self.set_xy(x+self.widths[i],y)
            self.line(x, y, x+self.widths[i], y)
            self.line(x, y, x, y+max_height)
        for x in a:
            self.line(x,y,x,y+max_height)
        self.ln(max_height)
        x=10
        self.line(x, self.get_y(), x+190, self.get_y())
    def footer(self):
        # Set position of the footer
        self.set_y(-10)
        # Set font size
        self.set_font('DejaVu', '', 8)
        # Add page number
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', 0, 0, 'C')
    def row(self, data):
        for row in data:
            max_height = 0
            a=[10, 35, 75,100,125,150,200]
            for i, item in enumerate(row):
                x = self.get_x()
                y = self.get_y()
                self.set_font('DejaVu',style="",size=11)
                self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
                if self.y - y > max_height:
                    max_height = self.y - y
                self.set_xy(x+self.widths[i],y)
                self.line(x, y, x+self.widths[i], y) #horizontal line
            self.ln(max_height)
            for x in a:
                self.line(x,y,x,y+max_height)
        x=10
        self.line(x, self.get_y(), x+190, self.get_y())
class ABC(FPDF):
    def __init__(self):
        super().__init__()
        # self.widths = [40, 60, 80]
        self.add_font('DejaVu','', r'ttf\DejaVuSansCondensed.ttf', uni=True)
        self.add_font('DejaVu','B', r'ttf\DejaVuSansCondensed-Bold.ttf', uni=True)
    def set_widths(self, widths):
        self.widths = widths
    def tieude(self,data):
        max_height = 0
        a=[10, 35, 75,105,135,165,200]
        for i, item in enumerate(data):
            x = self.get_x()
            y = self.get_y()
            self.set_font('DejaVu',style="B",size=11)
            self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
            if self.y - y > max_height:
                max_height = self.y - y
            self.set_xy(x+self.widths[i],y)
            self.line(x, y, x+self.widths[i], y)
            self.line(x, y, x, y+max_height)
        for x in a:
            self.line(x,y,x,y+max_height)
        self.ln(max_height)
        x=10
        self.line(x, self.get_y(), x+190, self.get_y())
    def footer(self):
        # Set position of the footer
        self.set_y(-10)
        # Set font size
        self.set_font('DejaVu', '', 8)
        # Add page number
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', 0, 0, 'C')
    def row(self, data):
        for row in data:
            max_height = 0
            a=[10, 35, 75,105,135,165,200]
            for i, item in enumerate(row):
                x = self.get_x()
                y = self.get_y()
                self.set_font('DejaVu',style="",size=11)
                self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
                if self.y - y > max_height:
                    max_height = self.y - y
                self.set_xy(x+self.widths[i],y)
                self.line(x, y, x+self.widths[i], y) #horizontal line
            self.ln(max_height)
            for x in a:
                self.line(x,y,x,y+max_height)
        x=10
        self.line(x, self.get_y(), x+190, self.get_y())     
class SS(FPDF):
    def __init__(self):
        super().__init__()
        # self.widths = [40, 60, 80]
        self.add_font('DejaVu','', r'ttf\DejaVuSansCondensed.ttf', uni=True)
        self.add_font('DejaVu','B', r'ttf\DejaVuSansCondensed-Bold.ttf', uni=True)
    def set_widths(self, widths):
        self.widths = widths
    def tieude(self,data):
        max_height = 0
        a=[10, 35, 90,130,170,200]
        for i, item in enumerate(data):
            x = self.get_x()
            y = self.get_y()
            self.set_font('DejaVu',style="B",size=11)
            self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
            if self.y - y > max_height:
                max_height = self.y - y
            self.set_xy(x+self.widths[i],y)
            self.line(x, y, x+self.widths[i], y)
            self.line(x, y, x, y+max_height)
        for x in a:
            self.line(x,y,x,y+max_height)
        self.ln(max_height)
        x=10
        self.line(x, self.get_y(), x+190, self.get_y())
    def footer(self):
        # Set position of the footer
        self.set_y(-10)
        # Set font size
        self.set_font('DejaVu', '', 8)
        # Add page number
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', 0, 0, 'C')
    def row(self, data):
        for row in data:
            max_height = 0
            a=[10, 35, 90,130,170,200]
            for i, item in enumerate(row):
                x = self.get_x()
                y = self.get_y()
                self.set_font('DejaVu',style="",size=11)
                self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
                if self.y - y > max_height:
                    max_height = self.y - y
                self.set_xy(x+self.widths[i],y)
                self.line(x, y, x+self.widths[i], y) #horizontal line
            self.ln(max_height)
            for x in a:
                self.line(x,y,x,y+max_height)
        x=10
        self.line(x, self.get_y(), x+190, self.get_y())  
class DatHang(FPDF):
    def __init__(self):
        super().__init__()
        # self.widths = [40, 60, 80]
        self.add_font('DejaVu','', r'ttf\DejaVuSansCondensed.ttf', uni=True)
        self.add_font('DejaVu','B', r'ttf\DejaVuSansCondensed-Bold.ttf', uni=True)
    def set_widths(self, widths):
        self.widths = widths
    def tieude(self,data):
        max_height = 0
        a=[10, 50, 80,140,160,200]
        for i, item in enumerate(data):
            x = self.get_x()
            y = self.get_y()
            self.set_font('DejaVu',style="B",size=11)
            self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
            if self.y - y > max_height:
                max_height = self.y - y
            self.set_xy(x+self.widths[i],y)
            self.line(x, y, x+self.widths[i], y)
            self.line(x, y, x, y+max_height)
        for x in a:
            self.line(x,y,x,y+max_height)
        self.ln(max_height)
        x=10
        self.line(x, self.get_y(), x+190, self.get_y())
    def footer(self):
        # Set position of the footer
        self.set_y(-10)
        # Set font size
        self.set_font('DejaVu', '', 8)
        # Add page number
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', 0, 0, 'C')
    def sum(self,data):
        self.widths=[150,40]
        for i, item in enumerate(data):
            x = self.get_x()
            y = self.get_y()
            self.set_font('DejaVu',style="B",size=11)
            self.multi_cell(self.widths[i], 7, str(item), border=1,align="C")
            self.set_xy(x+self.widths[i],y)
    def row(self, data):
        for row in data:
            max_height = 0
            a=[10, 50, 80,140,160,200]
            for i, item in enumerate(row):
                x = self.get_x()
                y = self.get_y()
                self.set_font('DejaVu',style="",size=11)
                self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
                if self.y - y > max_height:
                    max_height = self.y - y
                self.set_xy(x+self.widths[i],y)
                self.line(x, y, x+self.widths[i], y) #horizontal line
            self.ln(max_height)
            for x in a:
                self.line(x,y,x,y+max_height)
        x=10
        self.line(x, self.get_y(), x+190, self.get_y())  
class TK(FPDF):
    def __init__(self):
        super().__init__()
        # self.widths = [40, 60, 80]
        self.add_font('DejaVu','', r'ttf\DejaVuSansCondensed.ttf', uni=True)
        self.add_font('DejaVu','B', r'ttf\DejaVuSansCondensed-Bold.ttf', uni=True)
    def set_widths(self, widths):
        self.widths = widths
    def tieude(self,data):
        max_height = 0
        a=[20, 70, 120,190]
        for i, item in enumerate(data):
            x = self.get_x()
            y = self.get_y()
            self.set_font('DejaVu',style="B",size=11)
            self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
            if self.y - y > max_height:
                max_height = self.y - y
            self.set_xy(x+self.widths[i],y)
            self.line(x, y, x+self.widths[i], y)
            self.line(x, y, x, y+max_height)
        for x in a:
            self.line(x,y,x,y+max_height)
        self.ln(max_height)
        x=20
        self.line(x, self.get_y(), x+170, self.get_y())
    def footer(self):
        # Set position of the footer
        self.set_y(-10)
        # Set font size
        self.set_font('DejaVu', '', 8)
        # Add page number
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', 0, 0, 'C')
    def row(self, data):
        for row in data:
            max_height = 0
            a=[20, 70, 120,190]
            for i, item in enumerate(row):
                x = self.get_x()
                y = self.get_y()
                self.set_font('DejaVu',style="",size=11)
                self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
                if self.y - y > max_height:
                    max_height = self.y - y
                self.set_xy(x+self.widths[i],y)
                self.line(x, y, x+self.widths[i], y) #horizontal line
            self.ln(max_height)
            for x in a:
                self.line(x,y,x,y+max_height)
        x=20
        self.line(x, self.get_y(), x+170, self.get_y()) 
class PDFA010(FPDF):
    def __init__(self):
        super().__init__()
        # self.widths = [40, 60, 80]
        self.add_font('DejaVu','', r'ttf\DejaVuSansCondensed.ttf', uni=True)
        self.add_font('DejaVu','B', r'ttf\DejaVuSansCondensed-Bold.ttf', uni=True)
    def set_widths(self, widths):
        self.widths = widths
    def sum(self,data):
        self.widths=[140,50]
        for i, item in enumerate(data):
            x = self.get_x()
            y = self.get_y()
            self.set_font('DejaVu',style="B",size=11)
            self.multi_cell(self.widths[i], 7, str(item), border=1,align="C")
            self.set_xy(x+self.widths[i],y)
    def tieude(self,data):
        max_height = 0
        a=[10, 40, 80,150,200]
        for i, item in enumerate(data):
            x = self.get_x()
            y = self.get_y()
            self.set_font('DejaVu',style="B",size=11)
            self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
            if self.y - y > max_height:
                max_height = self.y - y
            self.set_xy(x+self.widths[i],y)
            self.line(x, y, x+self.widths[i], y)
            self.line(x, y, x, y+max_height)
        for x in a:
            self.line(x,y,x,y+max_height)
        self.ln(max_height)
        x=10
        self.line(x, self.get_y(), x+190, self.get_y())
    def footer(self):
        # Set position of the footer
        self.set_y(-10)
        # Set font size
        self.set_font('DejaVu', '', 8)
        # Add page number
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', 0, 0, 'C')
    def row(self, data):
        for row in data:
            max_height = 0
            a=[10, 40, 80,150,200]
            for i, item in enumerate(row):
                x = self.get_x()
                y = self.get_y()
                self.set_font('DejaVu',style="",size=11)
                self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
                if self.y - y > max_height:
                    max_height = self.y - y
                self.set_xy(x+self.widths[i],y)
                self.line(x, y, x+self.widths[i], y) #horizontal line
            self.ln(max_height)
            for x in a:
                self.line(x,y,x,y+max_height)
        x=10
        self.line(x, self.get_y(), x+190, self.get_y())
class PDFA040(FPDF):
    def __init__(self):
        super().__init__()
        # self.widths = [40, 60, 80]
        self.add_font('DejaVu','', r'ttf\DejaVuSansCondensed.ttf', uni=True)
        self.add_font('DejaVu','B', r'ttf\DejaVuSansCondensed-Bold.ttf', uni=True)
    def set_widths(self, widths):
        self.widths = widths
    def sum(self,data):
        self.widths=[165,25]
        for i, item in enumerate(data):
            x = self.get_x()
            y = self.get_y()
            self.set_font('DejaVu',style="B",size=11)
            self.multi_cell(self.widths[i], 7, str(item), border=1,align="C")
            self.set_xy(x+self.widths[i],y)
    def tieude(self,data):
        max_height = 0
        a=[10, 30,55, 115,175,200]
        for i, item in enumerate(data):
            x = self.get_x()
            y = self.get_y()
            self.set_font('DejaVu',style="B",size=11)
            self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
            if self.y - y > max_height:
                max_height = self.y - y
            self.set_xy(x+self.widths[i],y)
            self.line(x, y, x+self.widths[i], y)
            self.line(x, y, x, y+max_height)
        for x in a:
            self.line(x,y,x,y+max_height)
        self.ln(max_height)
        x=10
        self.line(x, self.get_y(), x+190, self.get_y())
    def footer(self):
        # Set position of the footer
        self.set_y(-10)
        # Set font size
        self.set_font('DejaVu', '', 8)
        # Add page number
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', 0, 0, 'C')
    def row(self, data):
        for row in data:
            max_height = 0
            a=[10, 30,55, 115,175,200]
            for i, item in enumerate(row):
                x = self.get_x()
                y = self.get_y()
                self.set_font('DejaVu',style="",size=11)
                self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
                if self.y - y > max_height:
                    max_height = self.y - y
                self.set_xy(x+self.widths[i],y)
                self.line(x, y, x+self.widths[i], y) #horizontal line
            self.ln(max_height)
            for x in a:
                self.line(x,y,x,y+max_height)
        x=10
        self.line(x, self.get_y(), x+190, self.get_y())
class PDFNhuCau(FPDF):
    def __init__(self):
        super().__init__()
        # self.widths = [40, 60, 80]
        self.add_font('DejaVu','', r'ttf\DejaVuSansCondensed.ttf', uni=True)
        self.add_font('DejaVu','B', r'ttf\DejaVuSansCondensed-Bold.ttf', uni=True)
    def set_widths(self, widths):
        self.widths = widths
    def sum(self,data):
        self.widths=[135,55]
        for i, item in enumerate(data):
            x = self.get_x()
            y = self.get_y()
            self.set_font('DejaVu',style="B",size=11)
            self.multi_cell(self.widths[i], 7, str(item), border=1,align="C")
            self.set_xy(x+self.widths[i],y)
    def tieude(self,data):
        max_height = 0
        a=[10, 40, 65,145,175,200]
        for i, item in enumerate(data):
            x = self.get_x()
            y = self.get_y()
            self.set_font('DejaVu',style="B",size=11)
            self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
            if self.y - y > max_height:
                max_height = self.y - y
            self.set_xy(x+self.widths[i],y)
            self.line(x, y, x+self.widths[i], y)
            self.line(x, y, x, y+max_height)
        for x in a:
            self.line(x,y,x,y+max_height)
        self.ln(max_height)
        x=10
        self.line(x, self.get_y(), x+190, self.get_y())
    def footer(self):
        # Set position of the footer
        self.set_y(-10)
        # Set font size
        self.set_font('DejaVu', '', 8)
        # Add page number
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', 0, 0, 'C')
    def row(self, data):
        for row in data:
            max_height = 0
            a=[10, 40, 65,145,175,200]
            for i, item in enumerate(row):
                x = self.get_x()
                y = self.get_y()
                self.set_font('DejaVu',style="",size=11)
                self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
                if self.y - y > max_height:
                    max_height = self.y - y
                self.set_xy(x+self.widths[i],y)
                self.line(x, y, x+self.widths[i], y) #horizontal line
            self.ln(max_height)
            for x in a:
                self.line(x,y,x,y+max_height)
        x=10
        self.line(x, self.get_y(), x+190, self.get_y())
class PDFDatHang(FPDF):
    def __init__(self):
        super().__init__()
        # self.widths = [40, 60, 80]
        self.add_font('DejaVu','', r'ttf\DejaVuSansCondensed.ttf', uni=True)
        self.add_font('DejaVu','B', r'ttf\DejaVuSansCondensed-Bold.ttf', uni=True)
    def set_widths(self, widths):
        self.widths = widths
    def sum(self,data):
        self.widths=[150,40]
        for i, item in enumerate(data):
            x = self.get_x()
            y = self.get_y()
            self.set_font('DejaVu',style="B",size=11)
            self.multi_cell(self.widths[i], 7, str(item), border=1,align="C")
            self.set_xy(x+self.widths[i],y)
    def tieude(self,data):
        max_height = 0
        a=[10, 60, 90,160,200]
        for i, item in enumerate(data):
            x = self.get_x()
            y = self.get_y()
            self.set_font('DejaVu',style="B",size=11)
            self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
            if self.y - y > max_height:
                max_height = self.y - y
            self.set_xy(x+self.widths[i],y)
            self.line(x, y, x+self.widths[i], y)
            self.line(x, y, x, y+max_height)
        for x in a:
            self.line(x,y,x,y+max_height)
        self.ln(max_height)
        x=10
        self.line(x, self.get_y(), x+190, self.get_y())
    def footer(self):
        # Set position of the footer
        self.set_y(-10)
        # Set font size
        self.set_font('DejaVu', '', 8)
        # Add page number
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', 0, 0, 'C')
    def row(self, data):
        for row in data:
            max_height = 0
            a=[10, 60, 90,160,200]
            for i, item in enumerate(row):
                x = self.get_x()
                y = self.get_y()
                self.set_font('DejaVu',style="",size=11)
                self.multi_cell(self.widths[i], 7, str(item), border=0,align="C")
                if self.y - y > max_height:
                    max_height = self.y - y
                self.set_xy(x+self.widths[i],y)
                self.line(x, y, x+self.widths[i], y) #horizontal line
            self.ln(max_height)
            for x in a:
                self.line(x,y,x,y+max_height)
        x=10
        self.line(x, self.get_y(), x+190, self.get_y())
class MplCanvas(FigureCanvasQTAgg):
    def __init__(self, parent=None, width=5, height=4, dpi=100):
        fig = Figure(figsize=(width, height), dpi=dpi)
        fig.patch.set_alpha(0)
        # fig.patch.set_facecolor('none')
        self.axes = fig.add_subplot(111)
        self.axes.set_facecolor('none')
        self.axes.spines['right'].set_visible(False)  # Xóa spine bên phải
        self.axes.spines['top'].set_visible(False) 
        # self.axes.spines['left'].set_visible(False) 
        self.axes.tick_params(axis='x', colors='white')  # Set the color of the x-axis ticks to white
        self.axes.tick_params(axis='y', colors='white',which='both', left=False, right=False, labelleft=False)
        super().__init__(fig)
        app_function = AppFunction()  # Instantiate the class containing the DataVisual function
        df = app_function.DataVisual(mpl_canvas=self)
        df1 = df.groupby(["TenVung"], as_index=False)["Giá trị tồn > 90d"].sum()
        df2= df.groupby(["TenVung"], as_index=False)["Giá trị tổng tồn"].sum()
        df=pd.merge(df1,df2,on='TenVung',how='inner')
        df['Giá trị tồn > 90d']=np.round(df['Giá trị tồn > 90d'],2)
        df['Giá trị tổng tồn']=np.round(df['Giá trị tổng tồn'],2)
        cot1_array = df['Giá trị tổng tồn'].values
        cot2_array = df['Giá trị tồn > 90d'].values
        clusters = ['Vùng 1', 'Vùng 2', 'Vùng 3']
        categories = ['Tổng tồn', 'Tồn > 90D']
        values = np.array([cot1_array, cot2_array]).T
        bar_width = 0.2
        x = np.arange(len(clusters))
        for i in range(len(categories)):
            self.axes.bar(x + (i * bar_width), values[:, i], bar_width, label=categories[i])
            for j, v in enumerate(values[:, i]):
                self.axes.text(x[j] + (i * bar_width), v, str(v), ha='center', va='bottom',color='white')
        self.axes.set_title('Tồn kho theo vùng',color='white')
        self.axes.set_xticks(x + ((len(categories) - 1) * bar_width) / 2)
        self.axes.set_xticklabels(clusters,color='white')
        self.figure.tight_layout()
class MplCanvasVung1(FigureCanvasQTAgg):
    def __init__(self, parent=None, width=5, height=4, dpi=100):
        fig = Figure(figsize=(width, height), dpi=dpi)
        fig.patch.set_alpha(0)
        # fig.patch.set_facecolor('none')
        self.axes = fig.add_subplot(111)
        self.axes.set_facecolor('none')
        self.axes.spines['right'].set_visible(False)  # Xóa spine bên phải
        self.axes.spines['top'].set_visible(False) 
        self.axes.spines['left'].set_visible(False) 
        self.axes.tick_params(axis='x', colors='white')  # Set the color of the x-axis ticks to white
        self.axes.tick_params(axis='y', colors='white',which='both', left=False, right=False, labelleft=False)
        super().__init__(fig)
        app_function = AppFunction()  # Instantiate the class containing the DataVisual function
        df = app_function.DataVisual(mpl_canvas=self)
        df1 = df.groupby(["TenVung"], as_index=False)["Giá trị tồn > 90d"].sum()
        df2= df.groupby(["TenVung"], as_index=False)["Giá trị tổng tồn"].sum()
        df=pd.merge(df1,df2,on=['TenVung'],how='inner')
        df=df.query('TenVung=="Vùng 1"')
        df['Giá trị tồn > 90d']=np.round(df['Giá trị tồn > 90d'],2)
        df['Giá trị tổng tồn']=np.round(df['Giá trị tổng tồn'],2)
        cot1_array = df['Giá trị tổng tồn'].values
        cot2_array = df['Giá trị tồn > 90d'].values
        clusters = ['Hà Nội']
        categories = ['Tổng tồn', 'Tồn > 90D']
        values = np.array([cot1_array, cot2_array]).T
        bar_width = 0.001
        x = np.arange(len(clusters))
        for i in range(len(categories)):
            self.axes.bar(x + (i * bar_width), values[:, i], bar_width, label=categories[i])
            for j, v in enumerate(values[:, i]):
                self.axes.text(x[j] + (i * bar_width), v, str(v), ha='center', va='bottom',color='white')
        self.axes.set_title('Vùng 1',color='white')
        self.axes.set_xticks(x + ((len(categories) - 1) * bar_width) / 2)
        self.axes.set_xticklabels(clusters,color='white')
        self.figure.tight_layout()
class MplCanvasVung2(FigureCanvasQTAgg):
    def __init__(self, parent=None, width=5, height=4, dpi=100):
        fig = Figure(figsize=(width, height), dpi=dpi)
        fig.patch.set_alpha(0)
        # fig.patch.set_facecolor('none')
        self.axes = fig.add_subplot(111)
        self.axes.set_facecolor('none')
        self.axes.spines['right'].set_visible(False)  # Xóa spine bên phải
        self.axes.spines['top'].set_visible(False) 
        self.axes.spines['left'].set_visible(False) 
        self.axes.tick_params(axis='x', colors='white')  # Set the color of the x-axis ticks to white
        self.axes.tick_params(axis='y', colors='white',which='both', left=False, right=False, labelleft=False)
        super().__init__(fig)
        app_function = AppFunction()  # Instantiate the class containing the DataVisual function
        df = app_function.DataVisual(mpl_canvas=self)
        df1 = df.groupby(["TenVung","TenVietTat"], as_index=False)["Giá trị tồn > 90d"].sum()
        df2= df.groupby(["TenVung","TenVietTat"], as_index=False)["Giá trị tổng tồn"].sum()
        df=pd.merge(df1,df2,on=['TenVung','TenVietTat'],how='inner')
        df=df.query('TenVung=="Vùng 2"')
        df = df.sort_values('Giá trị tổng tồn', ascending=False)
        df['Giá trị tồn > 90d']=np.round(df['Giá trị tồn > 90d'],2)
        df['Giá trị tổng tồn']=np.round(df['Giá trị tổng tồn'],2)
        cot1_array = df['Giá trị tổng tồn'].values
        cot2_array = df['Giá trị tồn > 90d'].values
        clusters = df['TenVietTat'].values
        categories = ['Tổng tồn', 'Tồn > 90D']
        values = np.array([cot1_array, cot2_array]).T
        bar_width = 0.2
        x = np.arange(len(clusters))
        for i in range(len(categories)):
            self.axes.bar(x + (i * bar_width), values[:, i], bar_width, label=categories[i])
            for j, v in enumerate(values[:, i]):
                self.axes.text(x[j] + (i * bar_width), v, str(v), ha='center', va='bottom',color='white')
        self.axes.set_title('Vùng 2',color='white')
        self.axes.set_xticks(x + ((len(categories) - 1) * bar_width) / 2)
        self.axes.set_xticklabels(clusters,color='white')
        self.figure.tight_layout()
class MplCanvasVung3(FigureCanvasQTAgg):
    def __init__(self, parent=None, width=5, height=4, dpi=100):
        fig = Figure(figsize=(width, height), dpi=dpi)
        fig.patch.set_alpha(0)
        # fig.patch.set_facecolor('none')
        self.axes = fig.add_subplot(111)
        self.axes.set_facecolor('none')
        self.axes.spines['right'].set_visible(False)  # Xóa spine bên phải
        self.axes.spines['top'].set_visible(False)
        self.axes.spines['left'].set_visible(False) 
        self.axes.tick_params(axis='x', colors='white')  # Set the color of the x-axis ticks to white
        self.axes.tick_params(axis='y', colors='white',which='both', left=False, right=False, labelleft=False)
        super().__init__(fig)
        app_function = AppFunction()  # Instantiate the class containing the DataVisual function
        df = app_function.DataVisual(mpl_canvas=self)
        df1 = df.groupby(["TenVung","TenVietTat"], as_index=False)["Giá trị tồn > 90d"].sum()
        df2= df.groupby(["TenVung","TenVietTat"], as_index=False)["Giá trị tổng tồn"].sum()
        df=pd.merge(df1,df2,on=['TenVung','TenVietTat'],how='inner')
        df=df.query('TenVung=="Vùng 3"')
        df = df.sort_values('Giá trị tổng tồn', ascending=False)
        df['Giá trị tồn > 90d']=np.round(df['Giá trị tồn > 90d'],2)
        df['Giá trị tổng tồn']=np.round(df['Giá trị tổng tồn'],2)
        cot1_array = df['Giá trị tổng tồn'].values
        cot2_array = df['Giá trị tồn > 90d'].values
        clusters = df['TenVietTat'].values
        categories = ['Tổng tồn', 'Tồn > 90D']
        values = np.array([cot1_array, cot2_array]).T
        bar_width = 0.2
        x = np.arange(len(clusters))
        for i in range(len(categories)):
            self.axes.bar(x + (i * bar_width), values[:, i], bar_width, label=categories[i])
            for j, v in enumerate(values[:, i]):
                self.axes.text(x[j] + (i * bar_width), v, str(v), ha='center', va='bottom',color='white')
        self.axes.set_title('Vùng 3',color='white')
        self.axes.set_xticks(x + ((len(categories) - 1) * bar_width) / 2)
        self.axes.set_xticklabels(clusters,color='white')
        self.figure.tight_layout()
class MplCanvasHMDT(FigureCanvasQTAgg):
    def __init__(self, parent=None, width=5, height=4, dpi=100):
        fig = Figure(figsize=(width, height), dpi=dpi)
        fig.patch.set_alpha(0)
        # fig.patch.set_facecolor('none')
        self.axes = fig.add_subplot(111)
        self.axes.set_facecolor('none')
        self.axes.spines['right'].set_visible(False)  # Xóa spine bên phải
        self.axes.spines['top'].set_visible(False) 
        self.axes.spines['right'].set_visible(False)  # Xóa spine bên phải
        self.axes.spines['top'].set_visible(False) 
        self.axes.tick_params(axis='x', colors='white')  # Set the color of the x-axis ticks to white
        self.axes.tick_params(axis='y', colors='white')
        super().__init__(fig)
        app_function = AppFunction()  # Instantiate the class containing the DataVisual function
        df = app_function.DataVisual(mpl_canvas=self)
        df1 = df.groupby(["TenHMDT"], as_index=False)["Giá trị tồn > 90d"].sum()
        df2= df.groupby(["TenHMDT"], as_index=False)["Giá trị tổng tồn"].sum()
        df=pd.merge(df1,df2,on=['TenHMDT'],how='inner')
        df['Giá trị tồn > 90d']=np.round(df['Giá trị tồn > 90d'],2)
        df['Giá trị tổng tồn']=np.round(df['Giá trị tổng tồn'],2)
        cot1_array = df['Giá trị tổng tồn'].values
        cot2_array = df['Giá trị tồn > 90d'].values
        clusters = df['TenHMDT'].values
        categories = ['Tổng tồn', 'Tồn > 90D']
        values = np.array([cot1_array, cot2_array]).T
        bar_width = 0.4
        y = np.arange(len(clusters))
        for i in range(len(categories)):
            self.axes.barh(y + (i * bar_width), values[:, i], bar_width, label=categories[i])
            for j, v in enumerate(values[:, i]):
                self.axes.text(v,y[j] + (i * bar_width), str(v), ha='left', va='center',color='white')
        self.axes.set_title('Tồn kho theo HMDT',color='white')
        self.axes.set_yticks(y + ((len(categories) - 1) * bar_width) / 2)
        self.axes.set_yticklabels(clusters,color='white')
        self.figure.tight_layout()
class MplCanvasNhomHang(FigureCanvasQTAgg):
    def __init__(self, parent=None, width=5, height=4, dpi=100):
        fig = Figure(figsize=(width, height), dpi=dpi)
        fig.patch.set_alpha(0)
        # fig.patch.set_facecolor('none')
        self.axes = fig.add_subplot(111)
        self.axes.set_facecolor('none')
        self.axes.spines['right'].set_visible(False)  # Xóa spine bên phải
        self.axes.spines['top'].set_visible(False) 
        self.axes.tick_params(axis='x', colors='white')  # Set the color of the x-axis ticks to white
        self.axes.tick_params(axis='y', colors='white')
        super().__init__(fig)
        app_function = AppFunction()  # Instantiate the class containing the DataVisual function
        df = app_function.DataVisual(mpl_canvas=self)
        df = df.groupby(["TenNhomNH"], as_index=False)["Giá trị tồn > 90d"].sum()
        df=df.nlargest(10,'Giá trị tồn > 90d')
        df = df.sort_values('Giá trị tồn > 90d', ascending=False)
        # df2= df.groupby(["TenNhomNH"], as_index=False)["Giá trị tổng tồn"].sum()
        # df=pd.merge(df1,df2,on=['TenNhomNH'],how='inner')
        df['Giá trị tồn > 90d']=np.round(df['Giá trị tồn > 90d'],2)
        # df['Giá trị tổng tồn']=np.round(df['Giá trị tổng tồn'],2)
        # cot1_array = df['Giá trị tổng tồn'].values
        cot2_array = df['Giá trị tồn > 90d'].values
        clusters = df['TenNhomNH'].values
        categories = ['Tồn > 90D']
        values = np.array([cot2_array]).T
        bar_width = 0.4
        y = np.arange(len(clusters))
        for i in range(len(categories)):
            self.axes.barh(y + (i * bar_width), values[:, i], bar_width, label=categories[i])
            for j, v in enumerate(values[:, i]):
                self.axes.text(v,y[j] + (i * bar_width), str(v), ha='left', va='center',color='white')
        self.axes.set_title('Top 10 MH tồn >90d',color='white')
        self.axes.set_yticks(y + ((len(categories) - 1) * bar_width) / 2)
        self.axes.set_yticklabels(clusters,color='white')
        self.figure.tight_layout()
class MplCanvasTyLe(FigureCanvasQTAgg):
    def __init__(self, parent=None, width=5, height=4, dpi=100):
        fig = Figure(figsize=(width, height), dpi=dpi)
        fig.patch.set_alpha(0)
        # fig.patch.set_facecolor('none')
        self.axes = fig.add_subplot(111)
        # self.axes.set_facecolor('#1f232a')
        self.axes.set_facecolor('none')
        self.axes.spines['right'].set_visible(False)  # Xóa spine bên phải
        self.axes.spines['top'].set_visible(False) 
        self.axes.tick_params(axis='x', colors='white')  # Set the color of the x-axis ticks to white
        self.axes.tick_params(axis='y', colors='white')
        super().__init__(fig)
        app_function = AppFunction()  # Instantiate the class containing the DataVisual function
        df = app_function.DataVisual(mpl_canvas=self)
        df1 = df.groupby(["TenVung"], as_index=False)["Giá trị tồn > 90d"].sum()
        df2= df.groupby(["TenVung"], as_index=False)["Giá trị tổng tồn"].sum()
        df=pd.merge(df1,df2,on=['TenVung'],how='inner')
        df = df.sort_values('Giá trị tổng tồn', ascending=False)
        df['Giá trị tồn > 90d']=np.round(df['Giá trị tồn > 90d'],2)
        df['Giá trị tổng tồn']=np.round(df['Giá trị tổng tồn'],2)
        cot1_array = df['Giá trị tổng tồn'].values
        cot2_array = df['Giá trị tồn > 90d'].values
        df['Tỷ lệ']=df['Giá trị tồn > 90d']/df['Giá trị tổng tồn']
        clusters = ['Vùng 1','Vùng 2','Vùng 3']
        categories = ['Tổng tồn', 'Tồn > 90D']
        values = np.array([cot1_array, cot2_array]).T
        bar_width = 0.2
        y = np.arange(len(clusters))
        for i in range(len(categories)):
            # self.axes.b
            self.axes.barh(y + (i * bar_width), values[:, i], bar_width, label=categories[i])
            for j, v in enumerate(values[:, i]):
                self.axes.text(v, y[j] + (i * bar_width), str(v), ha='left', va='center',color='white')
        self.axes.set_title('Tỷ lệ tồn 90D/Tổng tồn',color='white')
        self.axes.set_yticks(y + ((len(categories) - 1) * bar_width) / 2)
        self.axes.set_yticklabels(clusters,color='white')
        self.figure.tight_layout()
    # def tonTheoHMDT(self, df):
    #     # self.axes.barh(y='TenHMDT', width='Giá trị tồn khả dụng', data=df)
    #     cot1_array = df['Giá trị tổng tồn'].values
    #     cot2_array = df['Giá trị tồn > 90d'].values
    #     clusters = df['TenHMDT'].values
    #     categories = ['Tổng tồn', 'Tồn > 90D']
    #     values = np.array([cot1_array, cot2_array]).T
    #     bar_width = 0.2
    #     x = np.arange(len(clusters))
    #     for i in range(len(categories)):
    #         plt.bar(x + (i * bar_width), values[:, i], bar_width, label=categories[i])
    #     plt.title('Tồn kho theo HMDT')
    #     plt.xticks(x + ((len(categories) - 1) * bar_width) / 2, clusters)
    #     plt.show()
    #     self.ui.tonTheoHMDT.addWidget(plt)
    # def tonTheoNhomHang(self, df):
    #     self.axes.barh(y='TenNhomNH', width='Giá trị tồn > 90d', data=df)
    # def tyLeTon(self, df):
    #     df1 = df.groupby(["TenVung"], as_index=False)["Giá trị tồn > 90d"].sum()
    #     df2 = df.groupby(["TenVung"], as_index=False)["Giá trị tổng tồn"].sum()
    #     df=pd.merge(df1,df2,on='TenVung',how='inner')
    #     df['Tỷ lệ']=df['Giá trị tồn > 90d']/df['Giá trị tổng tồn']
    #     self.axes.barh(y='TenVung', width='Tỷ lệ', data=df)
    # def tonTheoVung(self, df):
    #     cot1_array = df['Giá trị tổng tồn'].values
    #     cot2_array = df['Giá trị tồn > 90d'].values
    #     clusters = ['Vùng 1', 'Vùng 2', 'Vùng 3']
    #     categories = ['Tổng tồn', 'Tồn > 90D']
    #     values = np.array([cot1_array, cot2_array]).T
    #     bar_width = 0.2
    #     x = np.arange(len(clusters))
    #     for i in range(len(categories)):
    #         plt.bar(x + (i * bar_width), values[:, i], bar_width, label=categories[i])
    #     plt.title('Tồn kho theo vùng')
    #     plt.xticks(x + ((len(categories) - 1) * bar_width) / 2, clusters)
    #     plt.show()
    # def tonVung1(self, df):
    #     df=df.query(df['TenVung']=='Vùng 1')
    #     cot1_array = df['Giá trị tổng tồn'].values
    #     cot2_array = df['Giá trị tồn > 90d'].values
    #     clusters = df['TenVietTat'].values
    #     categories = ['Tổng tồn', 'Tồn > 90d']
    #     values = np.array([cot1_array, cot2_array]).T
    #     bar_width = 0.2
    #     x = np.arange(len(clusters))
    #     for i in range(len(categories)):
    #         plt.bar(x + (i * bar_width), values[:, i], bar_width, label=categories[i])
    #     plt.title('Tồn kho theo vùng 1')
    #     plt.xticks(x + ((len(categories) - 1) * bar_width) / 2, clusters)
    #     plt.show()
    # def tonVung2(self, df):
    #     df=df.query(df['TenVung']=='Vùng 2')
    #     cot1_array = df['Giá trị tổng tồn'].values
    #     cot2_array = df['Giá trị tồn > 90d'].values
    #     clusters = df['TenVietTat'].values
    #     categories = ['Tổng tồn', 'Tồn > 90D']
    #     values = np.array([cot1_array, cot2_array]).T
    #     bar_width = 0.2
    #     x = np.arange(len(clusters))
    #     for i in range(len(categories)):
    #         plt.bar(x + (i * bar_width), values[:, i], bar_width, label=categories[i])
    #     plt.title('Tồn kho theo vùng 2')
    #     plt.xticks(x + ((len(categories) - 1) * bar_width) / 2, clusters)
    #     plt.show()
    # def tonVung3(self, df):
    #     df=df.query(df['TenVung']=='Vùng 3')
    #     cot1_array = df['Giá trị tổng tồn'].values
    #     cot2_array = df['Giá trị tồn > 90d'].values
    #     clusters = df['TenVietTat'].values
    #     categories = ['Tổng tồn', 'Tồn > 90D']
    #     values = np.array([cot1_array, cot2_array]).T
    #     bar_width = 0.2
    #     x = np.arange(len(clusters))
    #     for i in range(len(categories)):
    #         plt.bar(x + (i * bar_width), values[:, i], bar_width, label=categories[i])
    #     plt.title('Tồn kho theo vùng 3')
    #     plt.xticks(x + ((len(categories) - 1) * bar_width) / 2, clusters)
    #     plt.show()

class AlignDelegate(QtWidgets.QStyledItemDelegate):
    def initStyleOption(self, option, index):
        super(AlignDelegate, self).initStyleOption(option, index)
        option.displayAlignment = (QtCore.Qt.AlignCenter | QtCore.Qt.AlignRight)  
class AlignRDelegate(QtWidgets.QStyledItemDelegate):
    def initStyleOption(self, option, index):
        super(AlignRDelegate, self).initStyleOption(option, index)
        option.displayAlignment = (QtCore.Qt.AlignCenter)      


      
 

